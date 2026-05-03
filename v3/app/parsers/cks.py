"""ÇKS (.xlsx) parser."""
from __future__ import annotations
import io
import openpyxl
from app.utils import norm_koy


def parse_cks_xlsx(content: bytes, yil: int) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    data_start = 1
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
        if row[0] and str(row[0]).strip().upper() in ("İLÇESİ","İLÇE ADI","İLÇE"):
            data_start = i + 1; break
    rows = []
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        ilce = str(row[0] or "").strip(); koy = str(row[1] or "").strip(); sayi = row[2]
        if not ilce or not koy or sayi is None: continue
        try: sayi = int(float(sayi))
        except: continue
        rows.append({"yil":yil,"ilce":ilce.upper(),"koy":norm_koy(koy),"sayi":sayi})
    wb.close()
    return rows
