"""
Bitkisel üretim (.xlsx) parser.

Sütun konumları başlık satırından otomatik tespit edilir (import_router.py'deki
sabit indeksler yerine; farklı Excel formatlarını da kaldırır).
Geri dönüş değeri bulk_import.py ve import_router.py ile tam uyumludur.
"""
from __future__ import annotations

import io
import re
from datetime import datetime

import openpyxl

# Başlık metni → alan adı eşlemesi
_HEADER_MAP = {
    "İl":              "il",
    "İlçe":            "ilce",
    "Köy":             "koy",
    "Ürün":            "urun",
    "Tarım Şekli":     "tarim_sekli",
    "Tarım\nŞekli":    "tarim_sekli",
    "Ekili Alan":      "ekili_alan",   # "(da)" içerip içermediğine bakılmaz
    "Üretim Çeşidi":   "uretim_cesidi",
}


def _parse_yil(ws) -> int:
    """Başlık bölümünden 'Üretim Yılı: YYYY' yakala."""
    for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
        for cell in row:
            if cell and isinstance(cell, str) and "Üretim Yılı" in cell:
                m = re.search(r"(\d{4})", cell)
                if m:
                    return int(m.group(1))
    return datetime.now().year


def _find_header(ws) -> tuple[int, dict[str, int]]:
    """
    İlk 10 satırı tarayarak başlık satırını bulur.
    Döndürür: (veri_başlangıç_satırı, {alan_adı: 0-tabanlı_sütun_indeksi})
    """
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        cells = [str(c).strip() if c else "" for c in row]
        # 'İl', 'İlçe', 'Köy' üçü birden varsa başlık satırı
        if sum(1 for k in ("İl", "İlçe", "Köy") if any(k in c for c in cells)) == 3:
            col_map: dict[str, int] = {}
            for header, field in _HEADER_MAP.items():
                for ci, cell in enumerate(cells):
                    norm_h = header.replace("\n", "").replace(" ", "").lower()
                    norm_c = cell.replace("\n", "").replace(" ", "").lower()
                    if norm_h in norm_c:
                        col_map.setdefault(field, ci)
            return ri + 1, col_map
    # Geri dönüş: orijinal bulk_import.py'deki sabit indeksler
    return 7, {
        "il": 3, "ilce": 4, "koy": 5, "urun": 11,
        "tarim_sekli": 12, "ekili_alan": 16, "uretim_cesidi": 17,
    }


def parse_uretim_xlsx(content: bytes, yil: int) -> tuple[list[dict], str, int]:
    """
    Döndürür: (satırlar, ilce_adi, atlanan_sayı)
    Her satır dict: uretim_yili, il, ilce, koy, urun,
                    tarim_sekli, uretim_cesidi, ekili_alan
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    file_yil   = _parse_yil(ws)
    final_yil  = yil if yil else file_yil
    data_start, col_map = _find_header(ws)

    def get(row, field: str):
        i = col_map.get(field)
        return row[i] if i is not None and i < len(row) else None

    rows, ilce, skipped = [], "", 0
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        il_v    = get(row, "il")
        ilce_v  = get(row, "ilce")
        koy_v   = get(row, "koy")
        urun_v  = get(row, "urun")
        if not all([il_v, ilce_v, koy_v, urun_v]):
            skipped += 1
            continue
        try:
            alan = float(get(row, "ekili_alan") or 0)
        except (ValueError, TypeError):
            alan = 0.0
        if not ilce:
            ilce = str(ilce_v).strip().upper()
        rows.append({
            "uretim_yili":   final_yil,
            "il":            str(il_v).strip().upper(),
            "ilce":          str(ilce_v).strip().upper(),
            "koy":           str(koy_v).strip(),
            "urun":          str(urun_v).strip(),
            "tarim_sekli":   str(get(row, "tarim_sekli") or "Kuru").strip(),
            "uretim_cesidi": str(get(row, "uretim_cesidi") or "1.Üretim").strip(),
            "ekili_alan":    round(alan, 3),
        })

    wb.close()
    return rows, ilce, skipped
