"""
Excel/XLS parsing utilities for Burdur Tarım API.

Includes:
  - OLE/BIFF8 binary .xls reader (_read_xls_cells)
  - Individual parsers for each data type
"""
from __future__ import annotations

import io
import re
import struct as _s
from typing import Any

import openpyxl


# ═══════════════════════════════════════════════════════════
# LOW-LEVEL HELPERS
# ═══════════════════════════════════════════════════════════

def _clean(v: Any, maxlen: int = 200) -> str:
    """Sanitise a cell value to a safe string."""
    if v is None:
        return ""
    s = str(v).replace("\x00", "").strip()
    if any(ord(c) < 32 and c not in ("\t", "\n", "\r") for c in s):
        return ""
    return s[:maxlen]


def _num(v: Any) -> float:
    """Convert cell value to float, returning 0.0 on failure."""
    if v is None:
        return 0.0
    try:
        return round(float(v), 3)
    except (TypeError, ValueError):
        return 0.0


# ═══════════════════════════════════════════════════════════
# OLE / BIFF8 BINARY .XLS READER
# ═══════════════════════════════════════════════════════════

def _read_xls_cells(content: bytes) -> dict[tuple[int, int], Any]:
    """
    Minimal OLE/BIFF8 parser that extracts cell values from a legacy
    .xls file without any external library dependency.
    Returns a dict keyed by (row, col) → value.
    """
    SEC = 512

    # Build FAT from DIFAT
    difat = [s for s in _s.unpack_from("<109I", content, 76) if s < 0xFFFFFFFA]
    fat: list[int] = []
    for s in difat:
        fat.extend(_s.unpack_from(f"<{SEC // 4}I", content, 512 + s * SEC))

    # Locate Workbook stream in directory
    dir_off = 512 + _s.unpack_from("<I", content, 48)[0] * SEC
    wb_start = wb_size = 0
    for i in range(16):
        off = dir_off + i * 128
        nlen = _s.unpack_from("<H", content, off + 64)[0]
        if nlen < 2:
            continue
        name = content[off : off + nlen - 2].decode("utf-16-le", errors="replace")
        if "Workbook" in name or "workbook" in name:
            wb_start = _s.unpack_from("<I", content, off + 116)[0]
            wb_size = _s.unpack_from("<I", content, off + 120)[0]
            break
    if not wb_size:
        return {}

    # Read Workbook sectors
    wb = bytearray()
    sector, seen = wb_start, set()
    while sector < 0xFFFFFFFA and sector not in seen:
        seen.add(sector)
        wb.extend(content[512 + sector * SEC : 512 + sector * SEC + SEC])
        sector = fat[sector] if sector < len(fat) else 0xFFFFFFFF
    wb = bytes(wb[:wb_size])

    def _rk(rk: int) -> float:
        v = (
            (rk >> 2)
            if (rk & 2)
            else _s.unpack("<d", _s.pack("<Q", (rk & 0xFFFFFFFC) << 32))[0]
        )
        return v / 100.0 if (rk & 1) else v

    # Parse SST (Shared String Table)
    sst: list[str] = []
    pos = 0
    while pos < len(wb) - 4:
        rt = _s.unpack_from("<H", wb, pos)[0]
        rl = _s.unpack_from("<H", wb, pos + 2)[0]
        if rl > 65535:
            pos += 2
            continue
        body = wb[pos + 4 : pos + 4 + rl]
        if rt == 0x00FC:
            unique = _s.unpack_from("<I", body, 4)[0]
            sd = bytearray(body)
            np2 = pos + 4 + rl
            while np2 < len(wb) - 4:
                nrt = _s.unpack_from("<H", wb, np2)[0]
                nrl = _s.unpack_from("<H", wb, np2 + 2)[0]
                if nrt == 0x003C:
                    sd.extend(wb[np2 + 4 : np2 + 4 + nrl])
                    np2 += 4 + nrl
                else:
                    break
            p = 8
            for _ in range(unique):
                if p >= len(sd) - 2:
                    break
                try:
                    n = _s.unpack_from("<H", sd, p)[0]
                    fl = sd[p + 2]
                    p += 3
                    if fl & 0x08:
                        p += _s.unpack_from("<H", sd, p)[0] * 0 + 2
                    if fl & 0x04:
                        p += _s.unpack_from("<I", sd, p)[0] * 0 + 4
                    s = (
                        sd[p : p + n * 2].decode("utf-16-le", "replace")
                        if (fl & 1)
                        else sd[p : p + n].decode("cp1254", "replace")
                    )
                    p += n * 2 if (fl & 1) else n
                    sst.append(s.strip())
                except Exception:
                    p += 1
            break
        pos += 4 + rl

    # Parse cell records
    cells: dict[tuple[int, int], Any] = {}
    pos = 0
    while pos < len(wb) - 4:
        rt = _s.unpack_from("<H", wb, pos)[0]
        rl = _s.unpack_from("<H", wb, pos + 2)[0]
        if rl > 65535:
            pos += 2
            continue
        body = wb[pos + 4 : pos + 4 + rl]
        try:
            if rt == 0x00FD and rl >= 10:           # LabelSst
                r, col = _s.unpack_from("<HH", body)
                idx = _s.unpack_from("<I", body, 6)[0]
                if idx < len(sst):
                    cells[(r, col)] = sst[idx]
            elif rt == 0x0204 and rl >= 9:          # Label
                r, col = _s.unpack_from("<HH", body)
                n = _s.unpack_from("<H", body, 6)[0]
                fl = body[8]
                s = (
                    body[9 : 9 + n * 2].decode("utf-16-le", "replace")
                    if (fl & 1)
                    else body[9 : 9 + n].decode("cp1254", "replace")
                )
                cells[(r, col)] = s.strip()
            elif rt == 0x0203 and rl >= 14:         # Number
                r, col = _s.unpack_from("<HH", body)
                cells[(r, col)] = _s.unpack_from("<d", body, 6)[0]
            elif rt == 0x00BD and rl >= 6:          # MulRk
                r, fc = _s.unpack_from("<HH", body)
                lc = _s.unpack_from("<H", body, rl - 2)[0]
                for i, col in enumerate(range(fc, lc + 1)):
                    cells[(r, col)] = _rk(_s.unpack_from("<I", body, 4 + i * 6 + 2)[0])
            elif rt == 0x027E and rl >= 10:         # Rk
                r, col = _s.unpack_from("<HH", body)
                cells[(r, col)] = _rk(_s.unpack_from("<I", body, 6)[0])
        except Exception:
            pass
        pos += 4 + rl

    return cells


def _parse_yil_cols(
    cells: dict, start_col: int = 1, end_col: int = 10
) -> dict[int, int]:
    """Return {col_index: year} for numeric year headers in row 0."""
    yc: dict[int, int] = {}
    for col in range(start_col, end_col):
        v = cells.get((0, col))
        if v and str(v).strip().isdigit():
            yc[col] = int(float(str(v).strip()))
    return yc


# ═══════════════════════════════════════════════════════════
# BITKISEL ÜRETİM (ÇKS) — XLSX
# ═══════════════════════════════════════════════════════════

_HEADER_MAP = {
    "İl": "il",
    "İlçe": "ilce",
    "Köy": "koy",
    "Ürün": "urun",
    "Tarım Şekli": "tarim_sekli",
    "Ekili \nAlan (da)": "ekili_alan",
    "Üretim Çeşidi": "uretim_cesidi",
}


def find_uretim_columns(ws) -> tuple[int, dict[str, int]]:
    """
    Scan the first 10 rows for the header row that contains İl/İlçe/Köy.
    Returns (data_start_row, col_map) where col_map is {field: col_index}.
    """
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        rs = [str(c).strip() if c else "" for c in row]
        if "İl" in rs and "İlçe" in rs and "Köy" in rs:
            col_map: dict[str, int] = {}
            for header, field in _HEADER_MAP.items():
                norm_h = header.replace("\n", "").replace(" ", "")
                for i, cell in enumerate(rs):
                    if cell == header or norm_h in cell.replace("\n", "").replace(" ", ""):
                        col_map[field] = i
                        break
            return ri + 1, col_map
    return 7, {}


def parse_uretim_xlsx(
    content: bytes, yil: int
) -> tuple[list[dict], str, int]:
    """
    Parse a ÇKS bitkisel üretim XLSX file.
    Returns (rows, ilce_name, skipped_count).
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
    except Exception as exc:
        from app.exceptions import ExcelParseError
        raise ExcelParseError(str(exc))

    data_start, col_map = find_uretim_columns(ws)
    missing = {"il", "ilce", "koy", "urun", "ekili_alan"} - set(col_map.keys())
    if missing:
        wb.close()
        from fastapi import HTTPException
        raise HTTPException(422, f"Excel başlıkları bulunamadı: {missing}")

    rows: list[dict] = []
    ilce = ""
    skipped = 0

    for row in ws.iter_rows(min_row=data_start, values_only=True):
        def get(field: str):
            i = col_map.get(field)
            return row[i] if i is not None and i < len(row) else None

        il_v, ilce_v, koy_v, urun_v = get("il"), get("ilce"), get("koy"), get("urun")
        if not all([il_v, ilce_v, koy_v, urun_v]):
            skipped += 1
            continue

        if not ilce:
            ilce = str(ilce_v).strip().upper()

        try:
            alan = float(get("ekili_alan") or 0)
        except (TypeError, ValueError):
            alan = 0.0

        rows.append({
            "uretim_yili":   yil,
            "il":            str(il_v).strip().upper(),
            "ilce":          str(ilce_v).strip().upper(),
            "koy":           str(koy_v).strip(),
            "urun":          str(urun_v).strip(),
            "tarim_sekli":   str(get("tarim_sekli") or "Kuru").strip(),
            "uretim_cesidi": str(get("uretim_cesidi") or "1.Üretim").strip(),
            "ekili_alan":    round(alan, 3),
        })

    wb.close()
    return rows, ilce, skipped


# ═══════════════════════════════════════════════════════════
# HAYVANCILIK — XLS (binary)
# ═══════════════════════════════════════════════════════════

def parse_hayvancilik_xls(
    content: bytes, ilce_adi: str, yil: int
) -> list[dict]:
    cells = _read_xls_cells(content)
    if not cells:
        return []

    max_row = max(r for r, _ in cells)
    header_row = col_koy = col_sigir = col_manda = col_koyun = col_keci = None

    for r in range(min(25, max_row)):
        strs = {
            c: str(v)
            for c in range(25)
            if isinstance((v := cells.get((r, c), "")), str)
        }
        if any("Köy" in v and "Mahalle" in v for v in strs.values()) and any(
            kw in v for v in strs.values()
            for kw in ("Sığır", "Koyun", "Manda", "Keçi")
        ):
            header_row = r
            for col, v in strs.items():
                if "Köy" in v and "Mahalle" in v:
                    col_koy = col
                if "Sığır" in v:
                    col_sigir = col
                if "Manda" in v:
                    col_manda = col
                if "Koyun" in v:
                    col_koyun = col
                if "Keçi" in v:
                    col_keci = col
            break

    if header_row is None or col_koy is None:
        return []

    def num(r: int, col: int | None) -> int:
        if col is None:
            return 0
        try:
            return max(0, int(float(cells.get((r, col), 0) or 0)))
        except (TypeError, ValueError):
            return 0

    koyler: dict[str, dict] = {}
    for r in range(header_row + 1, max_row + 1):
        koy = cells.get((r, col_koy), "")
        if not isinstance(koy, str) or not koy.strip():
            continue
        koy = koy.strip().upper()
        si, ma, ko, ke = (
            num(r, col_sigir), num(r, col_manda),
            num(r, col_koyun), num(r, col_keci),
        )
        if koy not in koyler:
            koyler[koy] = dict(
                sigir=0, manda=0, koyun=0, keci=0,
                sigir_isletme=0, manda_isletme=0,
                koyun_isletme=0, keci_isletme=0,
                toplam_isletme=0,
            )
        k = koyler[koy]
        k["sigir"] += si
        k["manda"] += ma
        k["koyun"] += ko
        k["keci"] += ke
        if si > 0:
            k["sigir_isletme"] += 1
        if ma > 0:
            k["manda_isletme"] += 1
        if ko > 0:
            k["koyun_isletme"] += 1
        if ke > 0:
            k["keci_isletme"] += 1
        k["toplam_isletme"] += 1

    return [
        {"uretim_yili": yil, "il": "BURDUR", "ilce": ilce_adi.upper(), "koy": koy, **v}
        for koy, v in koyler.items()
    ]


# ═══════════════════════════════════════════════════════════
# KOOPERATİF — XLS (binary)
# ═══════════════════════════════════════════════════════════

def parse_kooperatif_xls(content: bytes) -> list[dict]:
    cells = _read_xls_cells(content)
    if not cells:
        return []

    max_row = max(r for r, _ in cells)
    rows: list[dict] = []

    for r in range(2, max_row + 1):
        ilce   = cells.get((r, 1))
        koy    = cells.get((r, 2))
        ktype  = cells.get((r, 3))
        ortak  = cells.get((r, 4))
        baskan = cells.get((r, 5))
        tel    = cells.get((r, 6))

        if not isinstance(ilce, str) or not _clean(ilce):
            continue
        if not isinstance(koy, str) or not _clean(koy):
            continue
        if not isinstance(ktype, str):
            continue

        tel_str = ""
        if tel:
            try:
                tel_str = str(int(float(tel)))
            except (TypeError, ValueError):
                tel_str = _clean(tel, 30)

        ortak_sayisi: int | None = None
        if ortak is not None:
            try:
                ortak_sayisi = int(float(ortak))
            except (TypeError, ValueError):
                pass

        rows.append({
            "ilce":          _clean(ilce, 60),
            "koy_belde":     _clean(koy, 120),
            "koop_turu":     _clean(ktype, 80),
            "ortak_sayisi":  ortak_sayisi,
            "baskan":        _clean(baskan, 200),
            "telefon":       tel_str,
        })

    return rows


# ═══════════════════════════════════════════════════════════
# SÜT DESTEKLEME — XLSX
# ═══════════════════════════════════════════════════════════

def parse_sut_xlsx(content: bytes, donem: str, yil: int) -> list[dict]:
    """
    Süt destekleme XLSX dosyasını parse eder.

    Başlık satırını otomatik bulur (ilk 20 satır içinde).
    Aranacak anahtar kelimeler:
      - Köy/Mahalle  → col_koy
      - İlçe         → col_ilce
      - İl           → col_il
      - Süt Miktarı / Temel Süt / Süt (lt) → col_sut
      - Destek Tutarı / Net Ödeme / Toplam → col_destek

    Başlık bulunamazsa sabit indekslere (eski davranış) geri düşer.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
    except Exception as exc:
        from app.exceptions import ExcelParseError
        raise ExcelParseError(str(exc))

    # ── 1. Başlık satırını otomatik bul ─────────────────────────────
    _SUT_KW    = ("süt miktarı", "temel süt", "süt (lt)", "süt lt", "sütmiktarı")
    _DESTEK_KW = ("destek tutarı", "net ödeme", "toplam tutar", "ödeme tutarı")
    _KOY_KW    = ("köy", "mahalle", "köy/mahalle")
    _ILCE_KW   = ("ilçe",)
    _IL_KW     = ("il",)

    def _match(cell_text: str, keywords: tuple) -> bool:
        t = cell_text.lower().replace(" ", "").replace("\n", "")
        return any(kw.replace(" ", "") in t for kw in keywords)

    header_row_idx = None
    col_koy = col_ilce = col_il = col_sut = col_destek = None

    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=0):
        cells_str = [str(c).strip() if c is not None else "" for c in row]
        # Köy + İlçe birlikte bulunuyorsa başlık satırı budur
        has_koy   = any(_match(c, _KOY_KW)    for c in cells_str)
        has_ilce  = any(_match(c, _ILCE_KW)   for c in cells_str)
        if has_koy and has_ilce:
            header_row_idx = ri
            for ci, c in enumerate(cells_str):
                if col_koy    is None and _match(c, _KOY_KW):    col_koy    = ci
                if col_ilce   is None and _match(c, _ILCE_KW):   col_ilce   = ci
                if col_il     is None and _match(c, _IL_KW) and not _match(c, _ILCE_KW):
                    col_il = ci
                if col_sut    is None and _match(c, _SUT_KW):    col_sut    = ci
                if col_destek is None and _match(c, _DESTEK_KW): col_destek = ci
            break

    # ── 2. Başlık bulunamadıysa eski sabit indekslere geri dön ──────
    if header_row_idx is None:
        col_koy    = 8
        col_ilce   = 9
        col_il     = 10
        col_sut    = 12
        col_destek = 25
        data_start = 11          # openpyxl min_row (1-tabanlı)
    else:
        data_start = header_row_idx + 2   # başlık satırının altındaki satırdan başla

    # ── 3. Satırları oku ─────────────────────────────────────────────
    def _safe(row, idx):
        """Kısa satırlarda IndexError vermez."""
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    rows: list[dict] = []
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if not row:
            continue

        # Tamamen boş satırları atla
        if all(c is None or str(c).strip() == "" for c in row):
            continue

        koy  = str(_safe(row, col_koy)  or "").strip().upper()
        ilce = str(_safe(row, col_ilce) or "").strip().upper()
        if not koy or not ilce:
            continue

        # Sayısal kontrol: süt veya destek sütunlarından en az biri dolu olmalı
        sut_val    = _safe(row, col_sut)
        destek_val = _safe(row, col_destek)
        has_number = isinstance(sut_val, (int, float)) or isinstance(destek_val, (int, float))
        if not has_number:
            continue

        rows.append({
            "donem":         donem,
            "yil":           yil,
            "il":            str(_safe(row, col_il) or "").strip().upper() or "BURDUR",
            "ilce":          ilce,
            "koy":           koy,
            "temel_sut_lt":  round(float(sut_val    or 0), 2),
            "destek_tutari": round(float(destek_val or 0), 2),
        })

    wb.close()
    return rows


# ═══════════════════════════════════════════════════════════
# ÇKS SAYISI — XLSX
# ═══════════════════════════════════════════════════════════

def _norm_koy(s: str) -> str:
    return re.sub(
        r"\s+(KÖYÜ?|MAHALLESİ?|BELDESİ?|MH\.?|KÖY\.?)$",
        "",
        s.strip().upper(),
    ).strip()


def parse_cks_xlsx(content: bytes, yil: int) -> list[dict]:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
    except Exception as exc:
        from app.exceptions import ExcelParseError
        raise ExcelParseError(str(exc))

    data_start = 1
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
        if row[0] and str(row[0]).strip().upper() in ("İLÇESİ", "İLÇE ADI", "İLÇE"):
            data_start = i + 1
            break

    rows: list[dict] = []
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        ilce = str(row[0] or "").strip()
        koy  = str(row[1] or "").strip()
        sayi = row[2]
        if not ilce or not koy or sayi is None:
            continue
        try:
            sayi = int(float(sayi))
        except (TypeError, ValueError):
            continue
        rows.append({
            "yil":  yil,
            "ilce": ilce.upper(),
            "koy":  _norm_koy(koy),
            "sayi": sayi,
        })

    wb.close()
    return rows


# ═══════════════════════════════════════════════════════════
# BİTKİSEL DESTEK — XLS (binary)
# ═══════════════════════════════════════════════════════════

def parse_bitkisel_xls(content: bytes, yil: int) -> list[dict]:
    cells = _read_xls_cells(content)
    if not cells:
        return []

    max_row = max(r for r, _ in cells)
    rows: list[dict] = []

    for r in range(4, max_row + 1):
        ilce = _clean(cells.get((r, 3)), 60)
        koy  = _clean(cells.get((r, 4)), 120)
        urun = _clean(cells.get((r, 10)), 120)
        if not ilce or not koy or not urun:
            continue
        rows.append({
            "yil":                  yil,
            "il":                   "BURDUR",
            "ilce":                 ilce.upper(),
            "koy":                  koy.upper(),
            "urun":                 urun.upper(),
            "feromon_adet":         _num(cells.get((r, 11))),
            "feromon_tuzak_adet":   _num(cells.get((r, 12))),
            "faydali_bocek_adet":   _num(cells.get((r, 13))),
            "desteklenen_alan_da":  _num(cells.get((r, 14))),
            "destek_tutari_tl":     _num(cells.get((r, 15))),
            "net_odeme_tl":         _num(cells.get((r, 17))),
        })

    return rows


# ═══════════════════════════════════════════════════════════
# ÖZET DESTEK TABLOLARI — XLS (binary)
# ═══════════════════════════════════════════════════════════

def parse_ozet_xls(
    content: bytes, key_field: str, start_col: int = 1
) -> list[dict]:
    cells = _read_xls_cells(content)
    if not cells:
        return []

    max_row = max(r for r, _ in cells)
    yil_cols = _parse_yil_cols(cells, start_col)
    if not yil_cols:
        return []

    rows: list[dict] = []
    for r in range(1, max_row + 1):
        ad = str(cells.get((r, 0), "") or "").strip()
        if not ad or any(t in ad.lower() for t in ("genel toplam", "toplam")):
            continue
        ad = re.sub(r"^[0-9]+-\s*", "", ad).strip()
        for col, yil in yil_cols.items():
            tutar = cells.get((r, col))
            try:
                tutar = round(float(tutar), 2) if tutar is not None else 0.0
            except (TypeError, ValueError):
                tutar = 0.0
            rows.append({key_field: ad, "yil": yil, "tutar_tl": tutar})

    return rows


def parse_fark_prim_xls(content: bytes) -> list[dict]:
    cells = _read_xls_cells(content)
    if not cells:
        return []

    max_row = max(r for r, _ in cells)
    yil_cols = _parse_yil_cols(cells, 2, 10)
    if not yil_cols:
        return []

    ana: dict[str, int] = {}
    for r in range(1, max_row + 1):
        a = str(cells.get((r, 0), "") or "").strip()
        b = str(cells.get((r, 1), "") or "").strip()
        if not a or "genel toplam" in a.lower():
            continue
        kat = re.sub(
            r"\s+Toplam.*$", "", a.replace("\n", " "), flags=re.IGNORECASE
        ).strip()
        if "toplam" in a.lower() or not b or b == a.split("\n")[0].strip():
            ana[kat] = r

    rows: list[dict] = []
    for kat, r in ana.items():
        for col, yil in yil_cols.items():
            tutar = cells.get((r, col))
            try:
                tutar = round(float(tutar), 2) if tutar is not None else 0.0
            except (TypeError, ValueError):
                tutar = 0.0
            rows.append({"kategori": kat, "yil": yil, "tutar_tl": tutar})

    return rows

# ═══════════════════════════════════════════════════════════
# PLANLI ÜRETİM DESTEĞİ — XLS (İCMAL-2, köy/mahalle detay)
# ═══════════════════════════════════════════════════════════

def _tr_norm(v: str) -> str:
    """Türkçe büyük harfleri de kapsayan normalize: boşluk/noktalama sil, küçük harf."""
    v = (v
         .replace("İ", "I").replace("Ş", "S").replace("Ğ", "G")
         .replace("Ü", "U").replace("Ö", "O").replace("Ç", "C")
         .upper()          # Tümünü ASCII büyük harfe çek
         .lower()          # Sonra küçük harfe
         .replace("ı", "i").replace("ş", "s").replace("ğ", "g")
         .replace("ü", "u").replace("ö", "o").replace("ç", "c")
         .replace(" ", "").replace("/", "").replace("\n", "")
         .replace("(", "").replace(")", "").replace("-", "")
         )
    return v


def parse_planli_uretim_xls(content: bytes, yil: int) -> tuple[list[dict], str, str]:
    """
    İCMAL-2 formatındaki Planlı Üretim Desteği XLS dosyasını parse eder.

    Başlık satırını otomatik bulur: 'İlçe' + 'Ürün/Ürün Grubu' + 'İşletme' içeren satır.
    İl/ilçe bilgisini metadata satırlarından ('İl: BURDUR', 'İlçe: BUCAK') okur.

    Returns (rows, il_adi, ilce_adi)
    """
    cells = _read_xls_cells(content)
    if not cells:
        return [], "", "", {}

    max_row = max(r for r, _ in cells)
    max_col = max(c for _, c in cells)

    # ── 1. Metadata'dan il/ilçe oku ─────────────────────────────────
    il_adi = "BURDUR"
    ilce_adi = ""
    for r in range(min(25, max_row)):
        for c in range(min(5, max_col + 1)):
            raw = str(cells.get((r, c), "") or "").strip()
            # Tüm Türkçe harfleri ASCII'ye normalize ederek karşılaştır
            ru = (raw
                  .replace("İ", "I").replace("ı", "I")
                  .replace("Ş", "S").replace("ş", "s")
                  .replace("Ç", "C").replace("ç", "c")
                  .replace("Ğ", "G").replace("ğ", "g")
                  .replace("Ü", "U").replace("ü", "u")
                  .replace("Ö", "O").replace("ö", "o")
                  .upper().replace(" ", ""))
            if ru.startswith("IL:") and "ILCE" not in ru:
                il_adi = raw.split(":", 1)[1].strip().upper()
            elif ru.startswith("ILCE:"):
                ilce_adi = raw.split(":", 1)[1].strip().upper()

    # ── 2. Başlık satırını bul ───────────────────────────────────────
    header_row = None
    col_map: dict[str, int] = {}

    for r in range(min(30, max_row)):
        row_vals = {c: str(cells.get((r, c), "") or "").strip()
                    for c in range(max_col + 1)}
        vn = {c: _tr_norm(v) for c, v in row_vals.items()}

        has_urun    = any("urun" in v for v in vn.values())
        has_ilce    = any(v == "ilce" for v in vn.values())
        has_isletme = any("isletme" in v for v in vn.values())

        if has_urun and has_ilce and has_isletme:
            header_row = r
            for c, v in vn.items():
                if v == "il":
                    col_map.setdefault("il", c)
                elif v == "ilce":
                    col_map.setdefault("ilce", c)
                elif "mahalle" in v or "koy" in v:
                    col_map.setdefault("koy", c)
                elif "urungrubu" in v or v == "urun":
                    col_map.setdefault("urun_grubu", c)
                elif "isletmesayisi" in v:
                    col_map.setdefault("isletme_sayisi", c)
                elif "yeralti" in v and "alan" in v:
                    col_map.setdefault("yeralti_su_alan_da", c)
                elif "desteklememiktari" in v or "desteklemiktari" in v:
                    col_map.setdefault("destekleme_miktari_tl", c)
                elif "destegetabi" in v and "alan" in v:
                    col_map.setdefault("destege_tabi_alan_da", c)
            break

    if header_row is None:
        return [], il_adi, ilce_adi, {}

    # ── 3. Satırları oku ─────────────────────────────────────────────
    rows: list[dict] = []
    toplam: dict = {}
    for r in range(header_row + 1, max_row + 1):
        il_val   = _clean(cells.get((r, col_map.get("il",   -1))), 60)  or il_adi
        ilce_val = _clean(cells.get((r, col_map.get("ilce", -1))), 60)  or ilce_adi
        koy_val  = _clean(cells.get((r, col_map.get("koy",  -1))), 120)
        urun_val = _clean(cells.get((r, col_map.get("urun_grubu", -1))), 120)

        if not koy_val:
            continue

        # TOPLAM satırını ayrıca yakala, kaydetme
        if any(kw in koy_val.lower() for kw in ("toplam", "genel")):
            def _tc(field: str, _r: int = r) -> float:
                return _num(cells.get((_r, col_map.get(field, -1))))
            def _ti(field: str, _r: int = r) -> int:
                v = cells.get((_r, col_map.get(field, -1)))
                try: return max(0, int(float(v or 0)))
                except: return 0
            toplam = {
                "isletme_sayisi":        _ti("isletme_sayisi"),
                "destege_tabi_alan_da":  _tc("destege_tabi_alan_da"),
                "yeralti_su_alan_da":    _tc("yeralti_su_alan_da"),
                "destekleme_miktari_tl": _tc("destekleme_miktari_tl"),
            }
            continue

        if not urun_val:
            continue
        if any(kw in urun_val.lower() for kw in ("toplam", "genel")):
            continue

        def _int_cell(field: str, _r: int = r) -> int:
            v = cells.get((_r, col_map.get(field, -1)))
            try:
                return max(0, int(float(v or 0)))
            except (TypeError, ValueError):
                return 0

        rows.append({
            "yil":                   yil,
            "il":                    il_val.upper(),
            "ilce":                  ilce_val.upper(),
            "koy":                   koy_val.upper(),
            "urun_grubu":            urun_val,
            "isletme_sayisi":        _int_cell("isletme_sayisi"),
            "destege_tabi_alan_da":  _num(cells.get((r, col_map.get("destege_tabi_alan_da", -1)))),
            "yeralti_su_alan_da":    _num(cells.get((r, col_map.get("yeralti_su_alan_da",  -1)))),
            "destekleme_miktari_tl": _num(cells.get((r, col_map.get("destekleme_miktari_tl", -1)))),
        })

    return rows, il_adi, ilce_adi, toplam


# ═══════════════════════════════════════════════════════════
# SERTİFİKALI FİDAN KULLANIM DESTEĞİ — XLS (İCMAL-2)
# ═══════════════════════════════════════════════════════════

def parse_sertifikali_fidan_xls(content: bytes, yil: int) -> tuple[list[dict], str, str]:
    """
    İCMAL-2 formatındaki Sertifikalı Fidan Kullanım Desteği XLS dosyasını parse eder.
    pandas bağımlılığı kaldırılmış — xlrd doğrudan kullanılır.

    Başlık satırını otomatik bulur: 'İlçe' + 'Fidan' + 'Kişi' içeren satır.
    İl/ilçe bilgisini metadata satırlarından ('İl : BURDUR', 'İlçe : BUCAK') okur.

    Sütunlar: Sıra No | İl | İlçe | Mahalle/Köy | Kişi Sayısı |
              Fidan Türü | Fidan Sayısı | Sertifikalı Alan (da) |
              Standart Alan (da) | Destekleme Alanı (da) | Destekleme Tutarı (TL)

    Returns (rows, il_adi, ilce_adi, toplam)
    """
    try:
        import xlrd as _xlrd
    except ImportError:
        return [], "", "", {}

    try:
        wb = _xlrd.open_workbook(file_contents=content)
    except Exception:
        return [], "", "", {}

    sh = wb.sheets()[0]
    nrows, ncols = sh.nrows, sh.ncols

    def _cell_str(r: int, c: int) -> str:
        """Hücreyi string olarak döndürür; xlrd'nin tırnaklı string sorununu düzeltir."""
        try:
            v = sh.cell_value(r, c)
            if v is None:
                return ""
            return str(v).strip().strip("'\"")
        except Exception:
            return ""

    def _cell_num(r: int, c: int) -> float:
        try:
            v = sh.cell_value(r, c)
            if sh.cell_type(r, c) == _xlrd.XL_CELL_NUMBER:
                return float(v)
            return float(str(v).strip().strip("'\""))
        except Exception:
            return 0.0

    def _int_cell(r: int, c: int) -> int:
        try:
            return max(0, int(_cell_num(r, c)))
        except Exception:
            return 0

    def _num_cell(r: int, c: int) -> float:
        return round(_cell_num(r, c), 3)

    # ── 1. Metadata: il / ilçe ─────────────────────────────────────
    il_adi = "BURDUR"
    ilce_adi = ""
    for r in range(min(25, nrows)):
        for c in range(ncols):
            raw = _cell_str(r, c)
            if not raw or ":" not in raw:
                continue
            norm = _tr_norm(raw)
            key = norm.partition(":")[0].strip()
            val = raw.split(":", 1)[1].strip().upper()
            if key == "il":
                il_adi = val
            elif key in ("ilce", "ilc"):
                ilce_adi = val

    # ── 2. Başlık satırını bul ──────────────────────────────────────
    header_row = None
    col_map: dict[str, int] = {}

    for r in range(min(25, nrows)):
        vn = {c: _tr_norm(_cell_str(r, c)) for c in range(ncols)}

        if (any("fidan" in v for v in vn.values()) and
                any(v == "ilce" for v in vn.values()) and
                any("kisi" in v for v in vn.values())):
            header_row = r
            for c, v in vn.items():
                if v == "il":
                    col_map.setdefault("il", c)
                elif v == "ilce":
                    col_map.setdefault("ilce", c)
                elif "mahalle" in v or "koy" in v:
                    col_map.setdefault("koy", c)
                elif "kisisayisi" in v:
                    col_map.setdefault("kisi_sayisi", c)
                elif "fidanturu" in v:
                    col_map.setdefault("fidan_turu", c)
                elif "fidansayisi" in v:
                    col_map.setdefault("fidan_sayisi", c)
                elif "sertifikali" in v and "alan" in v:
                    col_map.setdefault("sertifikali_alan_da", c)
                elif "standart" in v and "alan" in v:
                    col_map.setdefault("standart_alan_da", c)
                elif "destekleme" in v and "alan" in v and "tutari" not in v:
                    col_map.setdefault("destekleme_alani_da", c)
                elif "destekleme" in v and ("tutari" in v or "miktari" in v):
                    col_map.setdefault("destekleme_tutari_tl", c)
            break

    if header_row is None:
        return [], il_adi, ilce_adi, {}

    # ── 3. Veri satırlarını oku ────────────────────────────────────
    rows: list[dict] = []
    toplam: dict = {}
    for r in range(header_row + 1, nrows):
        koy_val   = _cell_str(r, col_map.get("koy",  -1))
        fidan_val = _cell_str(r, col_map.get("fidan_turu", -1))
        il_val    = _cell_str(r, col_map.get("il",   -1)) or il_adi
        ilce_val  = _cell_str(r, col_map.get("ilce", -1)) or ilce_adi

        if not koy_val:
            continue

        # TOPLAM satırını yakala, kaydetme
        if any(kw in koy_val.upper() for kw in ("TOPLAM", "GENEL")):
            toplam = {
                "kisi_sayisi":          _int_cell(r, col_map.get("kisi_sayisi",        -1)),
                "fidan_sayisi":         _int_cell(r, col_map.get("fidan_sayisi",        -1)),
                "sertifikali_alan_da":  _num_cell(r, col_map.get("sertifikali_alan_da", -1)),
                "standart_alan_da":     _num_cell(r, col_map.get("standart_alan_da",    -1)),
                "destekleme_alani_da":  _num_cell(r, col_map.get("destekleme_alani_da", -1)),
                "destekleme_tutari_tl": _num_cell(r, col_map.get("destekleme_tutari_tl",-1)),
            }
            continue

        rows.append({
            "yil":                  yil,
            "il":                   il_val.upper(),
            "ilce":                 ilce_val.upper(),
            "koy":                  koy_val.upper(),
            "fidan_turu":           fidan_val,
            "kisi_sayisi":          _int_cell(r, col_map.get("kisi_sayisi",        -1)),
            "fidan_sayisi":         _int_cell(r, col_map.get("fidan_sayisi",        -1)),
            "sertifikali_alan_da":  _num_cell(r, col_map.get("sertifikali_alan_da", -1)),
            "standart_alan_da":     _num_cell(r, col_map.get("standart_alan_da",    -1)),
            "destekleme_alani_da":  _num_cell(r, col_map.get("destekleme_alani_da", -1)),
            "destekleme_tutari_tl": _num_cell(r, col_map.get("destekleme_tutari_tl",-1)),
        })

    return rows, il_adi, ilce_adi, toplam


# ═══════════════════════════════════════════════════════════
# SERTİFİKALI TOHUM KULLANIM DESTEĞİ — XLS (İCMAL-2)
# ═══════════════════════════════════════════════════════════

def parse_sertifikali_tohum_xls(content: bytes, yil: int) -> tuple[list[dict], str, str, dict]:
    """
    İCMAL-2 formatındaki Sertifikalı Tohum Kullanım Desteği XLS dosyasını parse eder.

    Sütunlar: Sıra No | İl | İlçe | Mahalle/Köy | Ürün |
              İşletme Sayısı | Destekleme Alanı (da) | Destekleme Miktarı (TL)

    Returns (rows, il_adi, ilce_adi, toplam)
    """
    try:
        import xlrd as _xlrd
    except ImportError:
        return [], "", "", {}

    try:
        wb = _xlrd.open_workbook(file_contents=content)
    except Exception:
        return [], "", "", {}

    sh = wb.sheets()[0]
    nrows, ncols = sh.nrows, sh.ncols

    def _cell_str(r: int, c: int) -> str:
        try:
            v = sh.cell_value(r, c)
            if v is None:
                return ""
            return str(v).strip().strip("'\"")
        except Exception:
            return ""

    def _cell_num(r: int, c: int) -> float:
        try:
            v = sh.cell_value(r, c)
            if sh.cell_type(r, c) == _xlrd.XL_CELL_NUMBER:
                return float(v)
            return float(str(v).strip().strip("'\""))
        except Exception:
            return 0.0

    def _int_cell(r: int, c: int) -> int:
        try:
            return max(0, int(_cell_num(r, c)))
        except Exception:
            return 0

    def _num_cell(r: int, c: int) -> float:
        return round(_cell_num(r, c), 3)

    # ── 1. Metadata: il / ilçe ─────────────────────────────────────
    il_adi = "BURDUR"
    ilce_adi = ""
    for r in range(min(25, nrows)):
        for c in range(ncols):
            raw = _cell_str(r, c)
            if not raw or ":" not in raw:
                continue
            norm = _tr_norm(raw)
            key = norm.partition(":")[0].strip()
            val = raw.split(":", 1)[1].strip().upper()
            if key == "il":
                il_adi = val
            elif key in ("ilce", "ilc"):
                ilce_adi = val

    # ── 2. Başlık satırını bul ──────────────────────────────────────
    # Aranacak: 'ilce' + 'urun' + 'isletme' veya 'mahalle'/'koy' içeren satır
    header_row = None
    col_map: dict[str, int] = {}

    for r in range(min(25, nrows)):
        vn = {c: _tr_norm(_cell_str(r, c)) for c in range(ncols)}
        has_ilce    = any(v == "ilce" for v in vn.values())
        has_urun    = any("urun" in v for v in vn.values())
        has_isletme = any("isletme" in v for v in vn.values())

        if has_ilce and has_urun and has_isletme:
            header_row = r
            for c, v in vn.items():
                if v == "il":
                    col_map.setdefault("il", c)
                elif v == "ilce":
                    col_map.setdefault("ilce", c)
                elif "mahalle" in v or "koy" in v:
                    col_map.setdefault("koy", c)
                elif v == "urun" or "urunadi" in v:
                    col_map.setdefault("urun", c)
                elif "isletmesayisi" in v or v == "isletme":
                    col_map.setdefault("isletme_sayisi", c)
                elif "destekleme" in v and "alan" in v:
                    col_map.setdefault("destekleme_alani_da", c)
                elif "destekleme" in v and ("miktar" in v or "tutar" in v):
                    col_map.setdefault("destekleme_miktari_tl", c)
            break

    if header_row is None:
        return [], il_adi, ilce_adi, {}

    # ── 3. Veri satırlarını oku ────────────────────────────────────
    rows: list[dict] = []
    toplam: dict = {}

    for r in range(header_row + 1, nrows):
        koy_val  = _cell_str(r, col_map.get("koy",  -1))
        urun_val = _cell_str(r, col_map.get("urun", -1))
        il_val   = _cell_str(r, col_map.get("il",   -1)) or il_adi
        ilce_val = _cell_str(r, col_map.get("ilce", -1)) or ilce_adi

        if not koy_val:
            continue

        # TOPLAM satırını yakala, kaydetme
        if any(kw in koy_val.upper() for kw in ("TOPLAM", "GENEL")):
            toplam = {
                "isletme_sayisi":        _int_cell(r, col_map.get("isletme_sayisi",        -1)),
                "destekleme_alani_da":   _num_cell(r, col_map.get("destekleme_alani_da",   -1)),
                "destekleme_miktari_tl": _num_cell(r, col_map.get("destekleme_miktari_tl", -1)),
            }
            continue

        # Ürün sütunu boşsa bir önceki satırın ürününü kullanma — atla
        if not urun_val:
            # Bazen toplam satırında ürün C9'da olabilir
            urun_val = _cell_str(r, col_map.get("urun", -1))
        if not urun_val:
            continue

        if any(kw in urun_val.upper() for kw in ("TOPLAM", "GENEL")):
            continue

        rows.append({
            "yil":                  yil,
            "il":                   il_val.upper(),
            "ilce":                 ilce_val.upper(),
            "koy":                  koy_val.upper(),
            "urun":                 urun_val.upper(),
            "isletme_sayisi":       _int_cell(r, col_map.get("isletme_sayisi",        -1)),
            "destekleme_alani_da":  _num_cell(r, col_map.get("destekleme_alani_da",   -1)),
            "destekleme_miktari_tl":_num_cell(r, col_map.get("destekleme_miktari_tl", -1)),
        })

    return rows, il_adi, ilce_adi, toplam


# ═══════════════════════════════════════════════════════════
# TEMEL DESTEK — XLS (İCMAL-2, köy/mahalle detay)
# ═══════════════════════════════════════════════════════════

def parse_temel_destek_xls(content: bytes, yil: int) -> tuple[list[dict], str, str, dict]:
    """
    İCMAL-2 formatındaki Temel Destek XLS dosyasını parse eder.

    Sütunlar: Sıra No | İl | İlçe | Mahalle/Köy | Ürün/Ürün Grubu |
              İşletme Sayısı | Desteğe Tabi Alan (da) | Destekleme Miktarı (TL)

    Returns (rows, il_adi, ilce_adi, toplam)
    """
    try:
        import xlrd as _xlrd
    except ImportError:
        return [], "", "", {}

    try:
        wb = _xlrd.open_workbook(file_contents=content)
    except Exception:
        return [], "", "", {}

    sh = wb.sheets()[0]
    nrows, ncols = sh.nrows, sh.ncols

    def _cell_str(r: int, c: int) -> str:
        try:
            v = sh.cell_value(r, c)
            if v is None:
                return ""
            return str(v).strip().strip("'\"")
        except Exception:
            return ""

    def _cell_num(r: int, c: int) -> float:
        try:
            v = sh.cell_value(r, c)
            if sh.cell_type(r, c) == _xlrd.XL_CELL_NUMBER:
                return float(v)
            return float(str(v).strip().strip("'\""))
        except Exception:
            return 0.0

    def _int_cell(r: int, c: int) -> int:
        try:
            return max(0, int(_cell_num(r, c)))
        except Exception:
            return 0

    def _num_cell(r: int, c: int) -> float:
        return round(_cell_num(r, c), 3)

    # ── 1. Metadata: il / ilçe ─────────────────────────────────────
    il_adi = "BURDUR"
    ilce_adi = ""
    for r in range(min(25, nrows)):
        for c in range(ncols):
            raw = _cell_str(r, c)
            if not raw or ":" not in raw:
                continue
            norm = _tr_norm(raw)
            key = norm.partition(":")[0].strip()
            val = raw.split(":", 1)[1].strip().upper()
            if key == "il":
                il_adi = val
            elif key in ("ilce", "ilc"):
                ilce_adi = val

    # ── 2. Başlık satırını bul ──────────────────────────────────────
    # Aranacak: 'ilce' + 'urun' + 'isletme' içeren satır
    header_row = None
    col_map: dict[str, int] = {}

    for r in range(min(25, nrows)):
        vn = {c: _tr_norm(_cell_str(r, c)) for c in range(ncols)}
        has_ilce    = any(v == "ilce" for v in vn.values())
        has_urun    = any("urun" in v for v in vn.values())
        has_isletme = any("isletme" in v for v in vn.values())

        if has_ilce and has_urun and has_isletme:
            header_row = r
            for c, v in vn.items():
                if v == "il":
                    col_map.setdefault("il", c)
                elif v == "ilce":
                    col_map.setdefault("ilce", c)
                elif "mahalle" in v or "koy" in v:
                    col_map.setdefault("koy", c)
                elif "urun" in v:
                    col_map.setdefault("urun_grubu", c)
                elif "isletmesayisi" in v or v == "isletme":
                    col_map.setdefault("isletme_sayisi", c)
                elif "destegetabi" in v and "alan" in v:
                    col_map.setdefault("destege_tabi_alan_da", c)
                elif "destekleme" in v and ("miktar" in v or "tutar" in v):
                    col_map.setdefault("destekleme_miktari_tl", c)
            break

    if header_row is None:
        return [], il_adi, ilce_adi, {}

    # ── 3. Veri satırlarını oku ────────────────────────────────────
    rows: list[dict] = []
    toplam: dict = {}

    for r in range(header_row + 1, nrows):
        koy_val  = _cell_str(r, col_map.get("koy",  -1))
        urun_val = _cell_str(r, col_map.get("urun_grubu", -1))
        il_val   = _cell_str(r, col_map.get("il",   -1)) or il_adi
        ilce_val = _cell_str(r, col_map.get("ilce", -1)) or ilce_adi

        if not koy_val:
            continue

        # TOPLAM satırını yakala, kaydetme
        if any(kw in koy_val.upper() for kw in ("TOPLAM", "GENEL")):
            toplam = {
                "isletme_sayisi":       _int_cell(r, col_map.get("isletme_sayisi",       -1)),
                "destege_tabi_alan_da": _num_cell(r, col_map.get("destege_tabi_alan_da", -1)),
                "destekleme_miktari_tl":_num_cell(r, col_map.get("destekleme_miktari_tl",-1)),
            }
            continue

        if not urun_val or any(kw in urun_val.upper() for kw in ("TOPLAM", "GENEL")):
            continue

        rows.append({
            "yil":                   yil,
            "il":                    il_val.upper(),
            "ilce":                  ilce_val.upper(),
            "koy":                   koy_val.upper(),
            "urun_grubu":            urun_val,
            "isletme_sayisi":        _int_cell(r, col_map.get("isletme_sayisi",        -1)),
            "destege_tabi_alan_da":  _num_cell(r, col_map.get("destege_tabi_alan_da",  -1)),
            "destekleme_miktari_tl": _num_cell(r, col_map.get("destekleme_miktari_tl", -1)),
        })

    return rows, il_adi, ilce_adi, toplam


# ═══════════════════════════════════════════════════════════
# YEM BİTKİLERİ DESTEĞİ — XLS (İCMAL-2, köy/mahalle detay)
# ═══════════════════════════════════════════════════════════

def parse_yem_bitkileri_xls(content: bytes, yil: int) -> tuple[list[dict], str, str, dict]:
    """
    İCMAL-2 formatındaki Yem Bitkileri Desteği XLS dosyasını parse eder.

    Sütunlar: Sıra No | İl | İlçe | Mahalle/Köy | İşletme Sayısı | Ürün |
              Desteğe Tabi Alan (da) | Su Kısıtı | Süt Havzası | Destek Tutarı (TL)

    Returns (rows, il_adi, ilce_adi, toplam)
    """
    try:
        import xlrd as _xlrd
    except ImportError:
        return [], "", "", {}

    try:
        wb = _xlrd.open_workbook(file_contents=content)
    except Exception:
        return [], "", "", {}

    sh = wb.sheets()[0]
    nrows, ncols = sh.nrows, sh.ncols

    def _cell_str(r: int, c: int) -> str:
        try:
            v = sh.cell_value(r, c)
            if v is None:
                return ""
            return str(v).strip().strip("'\"")
        except Exception:
            return ""

    def _cell_num(r: int, c: int) -> float:
        try:
            v = sh.cell_value(r, c)
            if sh.cell_type(r, c) == _xlrd.XL_CELL_NUMBER:
                return float(v)
            return float(str(v).strip().strip("'\""))
        except Exception:
            return 0.0

    def _int_cell(r: int, c: int) -> int:
        try:
            return max(0, int(_cell_num(r, c)))
        except Exception:
            return 0

    def _num_cell(r: int, c: int) -> float:
        return round(_cell_num(r, c), 3)

    # ── 1. Metadata: il / ilçe ─────────────────────────────────────
    il_adi = "BURDUR"
    ilce_adi = ""
    for r in range(min(25, nrows)):
        for c in range(ncols):
            raw = _cell_str(r, c)
            if not raw or ":" not in raw:
                continue
            norm = _tr_norm(raw)
            key = norm.partition(":")[0].strip()
            val = raw.split(":", 1)[1].strip().upper()
            if key == "il":
                il_adi = val
            elif key in ("ilce", "ilc"):
                ilce_adi = val

    # ── 2. Başlık satırını bul ──────────────────────────────────────
    # Aranacak: 'ilce' + 'urun' + 'isletme' içeren satır
    header_row = None
    col_map: dict[str, int] = {}

    for r in range(min(25, nrows)):
        vn = {c: _tr_norm(_cell_str(r, c)) for c in range(ncols)}
        has_ilce    = any(v == "ilce" for v in vn.values())
        has_urun    = any(v == "urun" for v in vn.values())
        has_isletme = any("isletme" in v for v in vn.values())

        if has_ilce and has_urun and has_isletme:
            header_row = r
            for c, v in vn.items():
                if v == "il":
                    col_map.setdefault("il", c)
                elif v == "ilce":
                    col_map.setdefault("ilce", c)
                elif "mahalle" in v or "koy" in v:
                    col_map.setdefault("koy", c)
                elif v == "urun":
                    col_map.setdefault("urun", c)
                elif "isletmesayisi" in v or v == "isletme":
                    col_map.setdefault("isletme_sayisi", c)
                elif "destegetabi" in v and "alan" in v:
                    col_map.setdefault("destege_tabi_alan_da", c)
                elif "sukisiti" in v or "yeralti" in v:
                    col_map.setdefault("su_kisiti_da", c)
                elif "suthavzasi" in v:
                    col_map.setdefault("sut_havzasi_da", c)
                elif "destek" in v and ("tutar" in v or "miktar" in v):
                    col_map.setdefault("destek_tutari_tl", c)
            break

    if header_row is None:
        return [], il_adi, ilce_adi, {}

    # ── 3. Veri satırlarını oku ────────────────────────────────────
    rows: list[dict] = []
    toplam: dict = {}

    for r in range(header_row + 1, nrows):
        koy_val  = _cell_str(r, col_map.get("koy",  -1))
        urun_val = _cell_str(r, col_map.get("urun", -1))
        il_val   = _cell_str(r, col_map.get("il",   -1)) or il_adi
        ilce_val = _cell_str(r, col_map.get("ilce", -1)) or ilce_adi

        if not koy_val:
            continue

        # TOPLAM / GENEL TOPLAM satırını yakala
        if any(kw in koy_val.upper() for kw in ("TOPLAM", "GENEL")):
            toplam = {
                "isletme_sayisi":     _int_cell(r, col_map.get("isletme_sayisi",      -1)),
                "destege_tabi_alan_da": _num_cell(r, col_map.get("destege_tabi_alan_da", -1)),
                "su_kisiti_da":        _num_cell(r, col_map.get("su_kisiti_da",         -1)),
                "sut_havzasi_da":      _num_cell(r, col_map.get("sut_havzasi_da",       -1)),
                "destek_tutari_tl":    _num_cell(r, col_map.get("destek_tutari_tl",     -1)),
            }
            continue

        if not urun_val or any(kw in urun_val.upper() for kw in ("TOPLAM", "GENEL")):
            continue

        rows.append({
            "yil":                  yil,
            "il":                   il_val.upper(),
            "ilce":                 ilce_val.upper(),
            "koy":                  koy_val.upper(),
            "urun":                 urun_val.upper(),
            "isletme_sayisi":       _int_cell(r, col_map.get("isletme_sayisi",       -1)),
            "destege_tabi_alan_da": _num_cell(r, col_map.get("destege_tabi_alan_da", -1)),
            "su_kisiti_da":         _num_cell(r, col_map.get("su_kisiti_da",         -1)),
            "sut_havzasi_da":       _num_cell(r, col_map.get("sut_havzasi_da",       -1)),
            "destek_tutari_tl":     _num_cell(r, col_map.get("destek_tutari_tl",     -1)),
        })

    return rows, il_adi, ilce_adi, toplam


# ═══════════════════════════════════════════════════════════
# ZİRAİ DON DESTEĞİ — XLS (İCMAL-2, köy/mahalle detay)
# ═══════════════════════════════════════════════════════════

def parse_zirai_don_xls(content: bytes, yil: int) -> tuple[list[dict], str, str, dict]:
    """
    İCMAL-2 formatındaki Zirai Don Desteği XLS dosyasını parse eder.

    Sütunlar (geniş merge yapısı):
      Sıra No(C0) | İl(C3) | İlçe(C7) | Mahalle/Köy(C9) | Ürün(C13) |
      İşletme Sayısı(C16) | Hasar Oranı %(C18) | Birim Maliyet TL(C19) |
      Etkilenen Alan da(C22) | Toplam Masraf Tutarı TL(C26)

    Returns (rows, il_adi, ilce_adi, toplam)
    """
    try:
        import xlrd as _xlrd
    except ImportError:
        return [], "", "", {}

    try:
        wb = _xlrd.open_workbook(file_contents=content)
    except Exception:
        return [], "", "", {}

    sh = wb.sheets()[0]
    nrows, ncols = sh.nrows, sh.ncols

    def _cell_str(r: int, c: int) -> str:
        try:
            v = sh.cell_value(r, c)
            if v is None:
                return ""
            return str(v).strip().strip("'\"")
        except Exception:
            return ""

    def _cell_num(r: int, c: int) -> float:
        try:
            v = sh.cell_value(r, c)
            if sh.cell_type(r, c) == _xlrd.XL_CELL_NUMBER:
                return float(v)
            return float(str(v).strip().strip("'\""))
        except Exception:
            return 0.0

    def _int_cell(r: int, c: int) -> int:
        try:
            return max(0, int(_cell_num(r, c)))
        except Exception:
            return 0

    def _num_cell(r: int, c: int) -> float:
        return round(_cell_num(r, c), 3)

    # ── 1. Metadata: il / ilçe ─────────────────────────────────────
    il_adi = "BURDUR"
    ilce_adi = ""
    for r in range(min(20, nrows)):
        for c in range(ncols):
            raw = _cell_str(r, c)
            if not raw or ":" not in raw:
                continue
            norm = _tr_norm(raw)
            key = norm.partition(":")[0].strip()
            val = raw.split(":", 1)[1].strip().upper()
            if key == "il":
                il_adi = val
            elif key in ("ilce", "ilc"):
                ilce_adi = val

    # ── 2. Başlık satırını bul ──────────────────────────────────────
    # Aranacak: 'ilce' + 'urun' + 'isletme' + 'etkilenen' içeren satır
    header_row = None
    col_map: dict[str, int] = {}

    for r in range(min(20, nrows)):
        vn = {c: _tr_norm(_cell_str(r, c)) for c in range(ncols)}
        has_ilce      = any(v == "ilce" for v in vn.values())
        has_urun      = any(v == "urun" for v in vn.values())
        has_etkilenen = any("etkilenen" in v for v in vn.values())

        if has_ilce and has_urun and has_etkilenen:
            header_row = r
            for c, v in vn.items():
                if v == "il":
                    col_map.setdefault("il", c)
                elif v == "ilce":
                    col_map.setdefault("ilce", c)
                elif "mahalle" in v or "koy" in v:
                    col_map.setdefault("koy", c)
                elif v == "urun":
                    col_map.setdefault("urun", c)
                elif "isletmesayisi" in v:
                    col_map.setdefault("isletme_sayisi", c)
                elif "ortalamahasar" in v or "hasarorani" in v:
                    col_map.setdefault("hasar_orani_yuzde", c)
                elif "birimmaliyet" in v:
                    col_map.setdefault("birim_maliyet_tl", c)
                elif "etkilenen" in v and "alan" in v:
                    col_map.setdefault("etkilenen_alan_da", c)
                elif "toplammasraf" in v or ("toplam" in v and "tutar" in v):
                    col_map.setdefault("toplam_masraf_tl", c)
            break

    if header_row is None:
        return [], il_adi, ilce_adi, {}

    # ── 3. Veri satırlarını oku ────────────────────────────────────
    rows: list[dict] = []
    toplam: dict = {}

    for r in range(header_row + 1, nrows):
        koy_val  = _cell_str(r, col_map.get("koy",  -1))
        urun_val = _cell_str(r, col_map.get("urun", -1))
        il_val   = _cell_str(r, col_map.get("il",   -1)) or il_adi
        ilce_val = _cell_str(r, col_map.get("ilce", -1)) or ilce_adi

        if not koy_val:
            continue

        # TOPLAM satırı
        if any(kw in koy_val.upper() for kw in ("TOPLAM", "GENEL")):
            toplam = {
                "isletme_sayisi":    _int_cell(r, col_map.get("isletme_sayisi",    -1)),
                "etkilenen_alan_da": _num_cell(r, col_map.get("etkilenen_alan_da", -1)),
                "toplam_masraf_tl":  _num_cell(r, col_map.get("toplam_masraf_tl",  -1)),
            }
            continue

        if not urun_val or any(kw in urun_val.upper() for kw in ("TOPLAM", "GENEL")):
            continue

        rows.append({
            "yil":                yil,
            "il":                 il_val.upper(),
            "ilce":               ilce_val.upper(),
            "koy":                koy_val.upper(),
            "urun":               urun_val.upper(),
            "isletme_sayisi":     _int_cell(r, col_map.get("isletme_sayisi",    -1)),
            "hasar_orani_yuzde":  round(_cell_num(r, col_map.get("hasar_orani_yuzde", -1)), 2),
            "birim_maliyet_tl":   round(_cell_num(r, col_map.get("birim_maliyet_tl",  -1)), 2),
            "etkilenen_alan_da":  _num_cell(r, col_map.get("etkilenen_alan_da",  -1)),
            "toplam_masraf_tl":   round(_cell_num(r, col_map.get("toplam_masraf_tl",  -1)), 2),
        })

    return rows, il_adi, ilce_adi, toplam