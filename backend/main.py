"""
Burdur Tarım API
uvicorn main:app --reload --port 8000
"""
import io
import re
import time
import hashlib
import struct as _s
from contextlib import asynccontextmanager
from typing import Optional

from fastapi import FastAPI, Query, UploadFile, File, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy import create_engine, text
from dotenv import load_dotenv
import openpyxl
import os

load_dotenv()

DATABASE_URL = os.getenv("DATABASE_URL", "postgresql+psycopg2://postgres:postgres@localhost:5432/burdurdb")

engine = None

def init_engine():
    global engine
    engine = create_engine(DATABASE_URL, pool_size=5, max_overflow=10, pool_pre_ping=True)

# ═══════════════════════════════════════════════════════════
# TABLOLAR
# ═══════════════════════════════════════════════════════════
TABLES_SQL = [
    """CREATE TABLE IF NOT EXISTS uretim (
        id            SERIAL        PRIMARY KEY,
        uretim_yili   SMALLINT      NOT NULL,
        il            VARCHAR(60)   NOT NULL,
        ilce          VARCHAR(60)   NOT NULL,
        koy           VARCHAR(120)  NOT NULL,
        urun          VARCHAR(120)  NOT NULL,
        tarim_sekli   VARCHAR(20)   NOT NULL,
        uretim_cesidi VARCHAR(20)   NOT NULL,
        ekili_alan    NUMERIC(10,3) NOT NULL DEFAULT 0,
        created_at    TIMESTAMPTZ   NOT NULL DEFAULT NOW()
    )""",
    """CREATE TABLE IF NOT EXISTS hayvancilik (
        id             SERIAL       PRIMARY KEY,
        uretim_yili    SMALLINT     NOT NULL,
        il             VARCHAR(60)  NOT NULL,
        ilce           VARCHAR(60)  NOT NULL,
        koy            VARCHAR(120) NOT NULL,
        sigir          INTEGER      NOT NULL DEFAULT 0,
        manda          INTEGER      NOT NULL DEFAULT 0,
        koyun          INTEGER      NOT NULL DEFAULT 0,
        keci           INTEGER      NOT NULL DEFAULT 0,
        sigir_isletme  INTEGER      NOT NULL DEFAULT 0,
        manda_isletme  INTEGER      NOT NULL DEFAULT 0,
        koyun_isletme  INTEGER      NOT NULL DEFAULT 0,
        keci_isletme   INTEGER      NOT NULL DEFAULT 0,
        toplam_isletme INTEGER      NOT NULL DEFAULT 0,
        created_at     TIMESTAMPTZ  NOT NULL DEFAULT NOW()
    )""",
    """CREATE TABLE IF NOT EXISTS kooperatif (
        id           SERIAL       PRIMARY KEY,
        ilce         VARCHAR(60)  NOT NULL,
        koy_belde    VARCHAR(120) NOT NULL,
        koop_turu    VARCHAR(80)  NOT NULL,
        ortak_sayisi INTEGER,
        baskan       VARCHAR(200),
        telefon      VARCHAR(30),
        created_at   TIMESTAMPTZ  NOT NULL DEFAULT NOW()
    )""",
    """CREATE TABLE IF NOT EXISTS sut_destekleme (
        id            SERIAL        PRIMARY KEY,
        donem         VARCHAR(50)   NOT NULL,
        yil           SMALLINT      NOT NULL,
        il            VARCHAR(60)   NOT NULL,
        ilce          VARCHAR(60)   NOT NULL,
        koy           VARCHAR(120)  NOT NULL,
        temel_sut_lt  NUMERIC(14,2) NOT NULL DEFAULT 0,
        destek_tutari NUMERIC(14,2) NOT NULL DEFAULT 0,
        created_at    TIMESTAMPTZ   NOT NULL DEFAULT NOW()
    )""",
    """CREATE TABLE IF NOT EXISTS alan_bazli_destek (
        id         SERIAL        PRIMARY KEY,
        destek_adi VARCHAR(200)  NOT NULL,
        yil        SMALLINT      NOT NULL,
        tutar_tl   NUMERIC(16,2) NOT NULL DEFAULT 0,
        created_at TIMESTAMPTZ   NOT NULL DEFAULT NOW()
    )""",
    """CREATE TABLE IF NOT EXISTS fark_prim_destek (
        id         SERIAL        PRIMARY KEY,
        kategori   VARCHAR(200)  NOT NULL,
        yil        SMALLINT      NOT NULL,
        tutar_tl   NUMERIC(16,2) NOT NULL DEFAULT 0,
        created_at TIMESTAMPTZ   NOT NULL DEFAULT NOW()
    )""",
    """CREATE TABLE IF NOT EXISTS hayvancilik_destek (
        id         SERIAL        PRIMARY KEY,
        destek_adi VARCHAR(200)  NOT NULL,
        yil        SMALLINT      NOT NULL,
        tutar_tl   NUMERIC(16,2) NOT NULL DEFAULT 0,
        created_at TIMESTAMPTZ   NOT NULL DEFAULT NOW()
    )""",
    """CREATE TABLE IF NOT EXISTS genel_destek (
        id         SERIAL        PRIMARY KEY,
        destek_adi VARCHAR(200)  NOT NULL,
        yil        SMALLINT      NOT NULL,
        tutar_tl   NUMERIC(16,2) NOT NULL DEFAULT 0,
        created_at TIMESTAMPTZ   NOT NULL DEFAULT NOW()
    )""",
    """CREATE TABLE IF NOT EXISTS bitkisel_destek (
        id                  SERIAL        PRIMARY KEY,
        yil                 SMALLINT      NOT NULL,
        il                  VARCHAR(60)   NOT NULL DEFAULT 'BURDUR',
        ilce                VARCHAR(60)   NOT NULL,
        koy                 VARCHAR(120)  NOT NULL,
        urun                VARCHAR(120)  NOT NULL,
        feromon_adet        NUMERIC(12,2) NOT NULL DEFAULT 0,
        feromon_tuzak_adet  NUMERIC(12,2) NOT NULL DEFAULT 0,
        faydali_bocek_adet  NUMERIC(12,2) NOT NULL DEFAULT 0,
        desteklenen_alan_da NUMERIC(12,3) NOT NULL DEFAULT 0,
        destek_tutari_tl    NUMERIC(16,2) NOT NULL DEFAULT 0,
        net_odeme_tl        NUMERIC(16,2) NOT NULL DEFAULT 0,
        created_at          TIMESTAMPTZ   NOT NULL DEFAULT NOW(),
        updated_at          TIMESTAMPTZ   NOT NULL DEFAULT NOW()
    )""",
    """CREATE TABLE IF NOT EXISTS cks_sayisi (
        id   SERIAL       PRIMARY KEY,
        yil  SMALLINT     NOT NULL,
        ilce VARCHAR(60)  NOT NULL,
        koy  VARCHAR(120) NOT NULL,
        sayi INTEGER      NOT NULL DEFAULT 0
    )""",
    """CREATE TABLE IF NOT EXISTS import_log (
        id           SERIAL      PRIMARY KEY,
        dosya_adi    TEXT        NOT NULL,
        dosya_hash   VARCHAR(64),
        ilce         VARCHAR(60),
        uretim_yili  SMALLINT,
        kayit_sayisi INTEGER,
        silinen      INTEGER     DEFAULT 0,
        sure_sn      NUMERIC(8,2),
        durum        VARCHAR(20) NOT NULL DEFAULT 'basarili',
        hata_mesaji  TEXT,
        yuklendi_at  TIMESTAMPTZ NOT NULL DEFAULT NOW()
    )""",
]

INDEXES_SQL = [
    "CREATE INDEX IF NOT EXISTS idx_u_yil      ON uretim(uretim_yili)",
    "CREATE INDEX IF NOT EXISTS idx_u_ilce     ON uretim(ilce)",
    "CREATE INDEX IF NOT EXISTS idx_u_koy      ON uretim(koy)",
    "CREATE INDEX IF NOT EXISTS idx_u_urun     ON uretim(urun)",
    "CREATE INDEX IF NOT EXISTS idx_u_ilce_koy ON uretim(ilce, koy)",
    "CREATE INDEX IF NOT EXISTS idx_h_yil      ON hayvancilik(uretim_yili)",
    "CREATE INDEX IF NOT EXISTS idx_h_ilce     ON hayvancilik(ilce)",
    "CREATE INDEX IF NOT EXISTS idx_h_ilce_koy ON hayvancilik(ilce, koy)",
    "CREATE INDEX IF NOT EXISTS idx_k_ilce     ON kooperatif(ilce)",
    "CREATE INDEX IF NOT EXISTS idx_sd_yil     ON sut_destekleme(yil)",
    "CREATE INDEX IF NOT EXISTS idx_sd_ilce    ON sut_destekleme(ilce)",
    "CREATE INDEX IF NOT EXISTS idx_abd_yil    ON alan_bazli_destek(yil)",
    "CREATE INDEX IF NOT EXISTS idx_fpd_yil    ON fark_prim_destek(yil)",
    "CREATE INDEX IF NOT EXISTS idx_hd_yil     ON hayvancilik_destek(yil)",
    "CREATE INDEX IF NOT EXISTS idx_gd_yil     ON genel_destek(yil)",
    "CREATE INDEX IF NOT EXISTS idx_cks_yil    ON cks_sayisi(yil)",
    "CREATE INDEX IF NOT EXISTS idx_cks_ik     ON cks_sayisi(ilce, koy)",
    "CREATE UNIQUE INDEX IF NOT EXISTS idx_bd_unique ON bitkisel_destek(yil, ilce, koy, urun)",
    "CREATE INDEX IF NOT EXISTS idx_bd_yil     ON bitkisel_destek(yil)",
    "CREATE INDEX IF NOT EXISTS idx_bd_ilce    ON bitkisel_destek(ilce)",
    "CREATE INDEX IF NOT EXISTS idx_il_hash    ON import_log(dosya_hash)",
]

@asynccontextmanager
async def lifespan(app: FastAPI):
    init_engine()
    with engine.connect() as c:
        for sql in TABLES_SQL:
            c.execute(text(sql))
        for sql in INDEXES_SQL:
            c.execute(text(sql))
        c.commit()
    print("✅  DB hazır")
    yield
    engine.dispose()

# ═══════════════════════════════════════════════════════════
# APP
# ═══════════════════════════════════════════════════════════
app = FastAPI(title="Burdur Tarım API", lifespan=lifespan)
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_credentials=True, allow_methods=["*"], allow_headers=["*"])

# ═══════════════════════════════════════════════════════════
# YARDIMCILAR
# ═══════════════════════════════════════════════════════════
def _sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()

def _check_duplicate(conn, h: str) -> dict | None:
    row = conn.execute(text("SELECT dosya_adi, yuklendi_at FROM import_log WHERE dosya_hash=:h LIMIT 1"), {"h":h}).mappings().fetchone()
    return dict(row) if row else None

def _log_import(conn, dosya_adi, dosya_hash, yil, kayit, silinen, sure):
    conn.execute(text("INSERT INTO import_log (dosya_adi,dosya_hash,uretim_yili,kayit_sayisi,silinen,sure_sn,durum) VALUES (:d,:h,:y,:k,:s,:sn,'basarili')"),
                 {"d":dosya_adi,"h":dosya_hash,"y":yil,"k":kayit,"s":silinen,"sn":sure})

def _check_file(file: UploadFile, exts: set[str]):
    ext = (file.filename or "").rsplit(".",1)[-1].lower()
    if ext not in exts:
        raise HTTPException(400, f"Kabul edilen formatlar: {', '.join('.'+e for e in exts)}")

def _yil_from(yil_str: Optional[str], fname: str, default=2025) -> int:
    if yil_str and yil_str.isdigit(): return int(yil_str)
    m = re.search(r"(\d{4})", fname or "")
    return int(m.group(1)) if m else default

def _pagination(total: int, page: int, limit: int) -> dict:
    return {"total": total, "page": page, "limit": limit, "pages": max(1, -(-total // limit))}

def build_where(yil, ilce, koy, urun, tarim_sekli, uretim_cesidi, prefix=""):
    p = prefix + "." if prefix else ""
    clauses, params = [], {}
    if yil:           clauses.append(f"{p}uretim_yili=:yil");               params["yil"] = yil
    if ilce:          clauses.append(f"UPPER({p}ilce)=UPPER(:ilce)");        params["ilce"] = ilce
    if koy:           clauses.append(f"UPPER({p}koy)=UPPER(:koy)");          params["koy"] = koy
    if urun:          clauses.append(f"UPPER({p}urun) LIKE UPPER(:urun)");   params["urun"] = f"%{urun}%"
    if tarim_sekli:   clauses.append(f"{p}tarim_sekli=:tarim_sekli");        params["tarim_sekli"] = tarim_sekli
    if uretim_cesidi: clauses.append(f"{p}uretim_cesidi=:uretim_cesidi");    params["uretim_cesidi"] = uretim_cesidi
    return ("WHERE " + " AND ".join(clauses)) if clauses else "", params

# ═══════════════════════════════════════════════════════════
# OLE/BIFF8 XLS OKUYUCU
# ═══════════════════════════════════════════════════════════
def _read_xls_cells(content: bytes) -> dict:
    SEC = 512
    difat = [s for s in _s.unpack_from('<109I', content, 76) if s < 0xFFFFFFFA]
    fat: list[int] = []
    for s in difat:
        fat.extend(_s.unpack_from(f'<{SEC//4}I', content, 512+s*SEC))
    dir_off = 512 + _s.unpack_from('<I', content, 48)[0] * SEC
    wb_start = wb_size = 0
    for i in range(16):
        off = dir_off + i*128
        nlen = _s.unpack_from('<H', content, off+64)[0]
        if nlen < 2: continue
        name = content[off:off+nlen-2].decode('utf-16-le', errors='replace')
        if 'Workbook' in name or 'workbook' in name:
            wb_start = _s.unpack_from('<I', content, off+116)[0]
            wb_size  = _s.unpack_from('<I', content, off+120)[0]
            break
    if not wb_size: return {}
    wb = bytearray()
    sector, seen = wb_start, set()
    while sector < 0xFFFFFFFA and sector not in seen:
        seen.add(sector)
        wb.extend(content[512+sector*SEC:512+sector*SEC+SEC])
        sector = fat[sector] if sector < len(fat) else 0xFFFFFFFF
    wb = bytes(wb[:wb_size])

    def _rk(rk: int) -> float:
        v = (rk >> 2) if (rk & 2) else _s.unpack('<d', _s.pack('<Q', (rk & 0xFFFFFFFC) << 32))[0]
        return v / 100.0 if (rk & 1) else v

    sst: list[str] = []
    pos = 0
    while pos < len(wb) - 4:
        rt = _s.unpack_from('<H', wb, pos)[0]
        rl = _s.unpack_from('<H', wb, pos+2)[0]
        if rl > 65535: pos += 2; continue
        body = wb[pos+4:pos+4+rl]
        if rt == 0x00FC:
            unique = _s.unpack_from('<I', body, 4)[0]
            sd = bytearray(body)
            np2 = pos + 4 + rl
            while np2 < len(wb) - 4:
                nrt = _s.unpack_from('<H', wb, np2)[0]
                nrl = _s.unpack_from('<H', wb, np2+2)[0]
                if nrt == 0x003C: sd.extend(wb[np2+4:np2+4+nrl]); np2 += 4+nrl
                else: break
            p = 8
            for _ in range(unique):
                if p >= len(sd) - 2: break
                try:
                    n = _s.unpack_from('<H', sd, p)[0]; fl = sd[p+2]; p += 3
                    if fl & 0x08: p += _s.unpack_from('<H', sd, p)[0] * 0 + 2
                    if fl & 0x04: p += _s.unpack_from('<I', sd, p)[0] * 0 + 4
                    s = sd[p:p+n*2].decode('utf-16-le','replace') if (fl&1) else sd[p:p+n].decode('cp1254','replace')
                    p += n*2 if (fl&1) else n
                    sst.append(s.strip())
                except: p += 1
            break
        pos += 4 + rl

    cells: dict = {}
    pos = 0
    while pos < len(wb) - 4:
        rt = _s.unpack_from('<H', wb, pos)[0]
        rl = _s.unpack_from('<H', wb, pos+2)[0]
        if rl > 65535: pos += 2; continue
        body = wb[pos+4:pos+4+rl]
        try:
            if rt == 0x00FD and rl >= 10:
                r, col = _s.unpack_from('<HH', body)
                idx = _s.unpack_from('<I', body, 6)[0]
                if idx < len(sst): cells[(r,col)] = sst[idx]
            elif rt == 0x0204 and rl >= 9:
                r, col = _s.unpack_from('<HH', body)
                n = _s.unpack_from('<H', body, 6)[0]; fl = body[8]
                s = body[9:9+n*2].decode('utf-16-le','replace') if (fl&1) else body[9:9+n].decode('cp1254','replace')
                cells[(r,col)] = s.strip()
            elif rt == 0x0203 and rl >= 14:
                r, col = _s.unpack_from('<HH', body)
                cells[(r,col)] = _s.unpack_from('<d', body, 6)[0]
            elif rt == 0x00BD and rl >= 6:
                r, fc = _s.unpack_from('<HH', body)
                lc = _s.unpack_from('<H', body, rl-2)[0]
                for i, col in enumerate(range(fc, lc+1)):
                    cells[(r,col)] = _rk(_s.unpack_from('<I', body, 4+i*6+2)[0])
            elif rt == 0x027e and rl >= 10:
                r, col = _s.unpack_from('<HH', body)
                cells[(r,col)] = _rk(_s.unpack_from('<I', body, 6)[0])
        except: pass
        pos += 4 + rl
    return cells

def _clean(v, maxlen: int = 200) -> str:
    if v is None: return ""
    s = str(v).replace('\x00','').strip()
    if any(ord(c) < 32 and c not in ('\t','\n','\r') for c in s): return ""
    return s[:maxlen]

def _num(v) -> float:
    if v is None: return 0.0
    try: return round(float(v), 3)
    except: return 0.0

def _parse_yil_cols(cells: dict, sc=1, ec=10) -> dict:
    yc = {}
    for col in range(sc, ec):
        v = cells.get((0,col))
        if v and str(v).strip().isdigit(): yc[col] = int(float(str(v).strip()))
    return yc

# ═══════════════════════════════════════════════════════════
# BİTKİSEL ÜRETİM (ÇKS)
# ═══════════════════════════════════════════════════════════
HEADER_MAP = {"İl":"il","İlçe":"ilce","Köy":"koy","Ürün":"urun","Tarım Şekli":"tarim_sekli","Ekili \nAlan (da)":"ekili_alan","Üretim Çeşidi":"uretim_cesidi"}

def find_columns(ws) -> tuple[int, dict]:
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        rs = [str(c).strip() if c else "" for c in row]
        if "İl" in rs and "İlçe" in rs and "Köy" in rs:
            col_map = {}
            for th, f in HEADER_MAP.items():
                for i, cell in enumerate(rs):
                    if cell == th or th.replace("\n","").replace(" ","") in cell.replace("\n","").replace(" ",""):
                        col_map[f] = i; break
            return ri + 1, col_map
    return 7, {}

@app.get("/api/uretim")
def list_uretim(
    yil:           Optional[int] = Query(2025),
    ilce:          Optional[str] = Query(None),
    koy:           Optional[str] = Query(None),
    urun:          Optional[str] = Query(None),
    tarim_sekli:   Optional[str] = Query(None),
    uretim_cesidi: Optional[str] = Query(None),
    group_by:      Optional[str] = Query(None),
    sort_by:       Optional[str] = Query(None),
    sort_dir:      Optional[str] = Query("asc"),
    page:          int           = Query(1, ge=1),
    limit:         int           = Query(100, ge=1, le=50000),
):
    offset = (page-1)*limit
    where_u, params = build_where(yil, ilce, koy, urun, tarim_sekli, uretim_cesidi, prefix="u")
    where_p, _      = build_where(yil, ilce, koy, urun, tarim_sekli, uretim_cesidi)
    where = where_u if group_by in (None,"","urun_basit") else where_p
    if group_by not in (None,"","urun_basit"): params = _
    _ALL = {'uretim_yili','il','ilce','koy','urun','tarim_sekli','uretim_cesidi','ekili_alan','toplam_alan','kayit_sayisi','urun_cesidi','koy_sayisi','ilce_sayisi'}
    _def = 'toplam_alan' if group_by in ('koy','urun','urun_basit') else 'ilce'
    oc = sort_by if (sort_by and sort_by in _ALL) else _def
    od = 'DESC' if str(sort_dir or '').lower()=='desc' else 'ASC'
    if group_by == "koy":
        sql = f"SELECT ilce,koy,COUNT(DISTINCT urun)::int AS urun_cesidi,COUNT(*)::int AS kayit_sayisi,ROUND(SUM(ekili_alan)::numeric,2) AS toplam_alan FROM uretim {where} GROUP BY ilce,koy ORDER BY {oc} {od} LIMIT :limit OFFSET :offset"
        cnt = f"SELECT COUNT(*) FROM (SELECT koy FROM uretim {where} GROUP BY ilce,koy) s"
    elif group_by == "urun":
        sql = f"SELECT urun,tarim_sekli,COUNT(DISTINCT ilce)::int AS ilce_sayisi,COUNT(DISTINCT koy)::int AS koy_sayisi,COUNT(*)::int AS kayit_sayisi,ROUND(SUM(ekili_alan)::numeric,2) AS toplam_alan FROM uretim {where} GROUP BY urun,tarim_sekli ORDER BY {oc} {od} LIMIT :limit OFFSET :offset"
        cnt = f"SELECT COUNT(*) FROM (SELECT urun FROM uretim {where} GROUP BY urun,tarim_sekli) s"
    elif group_by == "urun_basit":
        sql = f"SELECT u.ilce,u.koy,REGEXP_REPLACE(u.urun,' *[(][^)]*[)]','','g') AS urun,ROUND(SUM(u.ekili_alan)::numeric,2) AS toplam_alan,MAX(cs.sayi) AS ciftci_sayisi FROM uretim u LEFT JOIN cks_sayisi cs ON UPPER(cs.ilce)=UPPER(u.ilce) AND UPPER(cs.koy)=UPPER(u.koy) AND cs.yil=u.uretim_yili {where} GROUP BY u.ilce,u.koy,REGEXP_REPLACE(u.urun,' *[(][^)]*[)]','','g') ORDER BY {oc} {od} LIMIT :limit OFFSET :offset"
        cnt = f"SELECT COUNT(*) FROM (SELECT 1 FROM uretim u LEFT JOIN cks_sayisi cs ON UPPER(cs.ilce)=UPPER(u.ilce) AND UPPER(cs.koy)=UPPER(u.koy) AND cs.yil=u.uretim_yili {where} GROUP BY u.ilce,u.koy,REGEXP_REPLACE(u.urun,' *[(][^)]*[)]','','g')) s"
    else:
        sql = f"SELECT u.uretim_yili,u.il,u.ilce,u.koy,u.urun,u.tarim_sekli,u.uretim_cesidi,ROUND(SUM(u.ekili_alan)::numeric,3) AS ekili_alan,MAX(cs.sayi) AS ciftci_sayisi FROM uretim u LEFT JOIN cks_sayisi cs ON UPPER(cs.ilce)=UPPER(u.ilce) AND UPPER(cs.koy)=UPPER(u.koy) AND cs.yil=u.uretim_yili {where} GROUP BY u.uretim_yili,u.il,u.ilce,u.koy,u.urun,u.tarim_sekli,u.uretim_cesidi ORDER BY {oc} {od} LIMIT :limit OFFSET :offset"
        cnt = f"SELECT COUNT(*) FROM (SELECT 1 FROM uretim u LEFT JOIN cks_sayisi cs ON UPPER(cs.ilce)=UPPER(u.ilce) AND UPPER(cs.koy)=UPPER(u.koy) AND cs.yil=u.uretim_yili {where} GROUP BY u.uretim_yili,u.il,u.ilce,u.koy,u.urun,u.tarim_sekli,u.uretim_cesidi) s"
    with engine.connect() as conn:
        rows  = conn.execute(text(sql), {**params,"limit":limit,"offset":offset}).mappings().all()
        total = int(conn.execute(text(cnt), params).scalar() or 0)
    return {"data":[dict(r) for r in rows], **_pagination(total,page,limit)}


@app.get("/api/uretim/ozet")
def ozet_uretim(yil: int = Query(2025), ilce: Optional[str] = Query(None), group_by: str = Query("ilce"), limit: int = Query(50, le=200)):
    where, params = build_where(yil, ilce, None, None, None, None)
    VALID = {"ilce","koy","urun","tarim_sekli","uretim_cesidi"}
    grp = group_by if group_by in VALID else "ilce"
    EXTRA = {"ilce":"COUNT(DISTINCT koy)::int AS koy_sayisi,COUNT(DISTINCT urun)::int AS urun_cesidi,","koy":"ilce,COUNT(DISTINCT urun)::int AS urun_cesidi,","urun":"tarim_sekli,COUNT(DISTINCT ilce)::int AS ilce_sayisi,COUNT(DISTINCT koy)::int AS koy_sayisi,","tarim_sekli":"COUNT(DISTINCT urun)::int AS urun_cesidi,","uretim_cesidi":"COUNT(DISTINCT urun)::int AS urun_cesidi,"}
    EGRP = {"koy":", ilce","urun":", tarim_sekli"}
    sql = f"SELECT {grp},{EXTRA.get(grp,'')} COUNT(*)::int AS kayit_sayisi,ROUND(SUM(ekili_alan)::numeric,2) AS toplam_alan_da,ROUND((SUM(ekili_alan)/10.0)::numeric,3) AS toplam_alan_ha FROM uretim {where} GROUP BY {grp}{EGRP.get(grp,'')} ORDER BY toplam_alan_da DESC LIMIT :limit"
    tot_sql = f"SELECT ROUND(SUM(ekili_alan)::numeric,2),COUNT(*) FROM uretim {where}"
    with engine.connect() as c:
        rows = c.execute(text(sql), {**params,"limit":limit}).mappings().all()
        tot  = c.execute(text(tot_sql), params).fetchone()
    return {"group_by":grp,"data":[dict(r) for r in rows],"toplam_alan_da":float(tot[0] or 0),"toplam_kayit":int(tot[1] or 0)}


@app.get("/api/uretim/urunler")
def list_urunler(yil: int = Query(2025), ilce: Optional[str] = Query(None)):
    where, params = build_where(yil, ilce, None, None, None, None)
    with engine.connect() as c:
        rows = c.execute(text(f"SELECT urun,ROUND(SUM(ekili_alan)::numeric,2) AS toplam_alan FROM uretim {where} GROUP BY urun ORDER BY toplam_alan DESC"), params).mappings().all()
    return {"data":[dict(r) for r in rows]}


@app.get("/api/uretim/ilceler")
def list_ilceler(yil: int = Query(2025)):
    with engine.connect() as c:
        rows = c.execute(text("SELECT ilce,COUNT(DISTINCT koy)::int AS koy_sayisi,ROUND(SUM(ekili_alan)::numeric,2) AS toplam_alan FROM uretim WHERE uretim_yili=:yil GROUP BY ilce ORDER BY ilce"), {"yil":yil}).mappings().all()
    return {"data":[dict(r) for r in rows]}


@app.get("/api/uretim/log")
def import_log_list(limit: int = Query(20, le=100)):
    with engine.connect() as c:
        rows = c.execute(text("SELECT * FROM import_log ORDER BY yuklendi_at DESC LIMIT :limit"), {"limit":limit}).mappings().all()
    return {"data":[dict(r) for r in rows]}


@app.delete("/api/uretim/temizle")
def uretim_temizle(yil: Optional[int] = Query(None), ilce: Optional[str] = Query(None)):
    with engine.begin() as c:
        if yil and ilce: r = c.execute(text("DELETE FROM uretim WHERE uretim_yili=:y AND ilce=:i"), {"y":yil,"i":ilce.upper()})
        elif yil:        r = c.execute(text("DELETE FROM uretim WHERE uretim_yili=:y"), {"y":yil})
        else:            r = c.execute(text("DELETE FROM uretim"))
    return {"silinen":r.rowcount}


@app.post("/api/import")
async def import_excel(file: UploadFile = File(...), yil: Optional[str] = Form(None), truncate: Optional[str] = Form("true")):
    _check_file(file, {"xlsx","xls","xlsm","ods"})
    content = await file.read()
    fh = _sha256(content)
    with engine.connect() as _c:
        dup = _check_duplicate(_c, fh)
    if dup: raise HTTPException(409, f"Bu dosya daha önce yüklenmiş ({dup['dosya_adi']})")
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(422, f"Excel okunamadı: {e}")
    final_yil = _yil_from(yil, file.filename or "")
    data_start, col_map = find_columns(ws)
    missing = {"il","ilce","koy","urun","ekili_alan"} - set(col_map.keys())
    if missing: raise HTTPException(422, f"Excel'de başlıklar bulunamadı: {missing}")
    rows, ilce, skipped = [], "", 0
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        def get(f): i=col_map.get(f); return row[i] if i is not None and i<len(row) else None
        il_v,ilce_v,koy_v,urun_v = get("il"),get("ilce"),get("koy"),get("urun")
        if not all([il_v,ilce_v,koy_v,urun_v]): skipped+=1; continue
        if not ilce: ilce = str(ilce_v).strip().upper()
        try: alan = float(get("ekili_alan") or 0)
        except: alan = 0.0
        rows.append({"uretim_yili":final_yil,"il":str(il_v).strip().upper(),"ilce":str(ilce_v).strip().upper(),"koy":str(koy_v).strip(),"urun":str(urun_v).strip(),"tarim_sekli":str(get("tarim_sekli") or "Kuru").strip(),"uretim_cesidi":str(get("uretim_cesidi") or "1.Üretim").strip(),"ekili_alan":round(alan,3)})
    wb.close()
    if not rows: raise HTTPException(422, f"Geçerli veri bulunamadı (atlanan: {skipped})")
    t0 = time.perf_counter()
    with engine.begin() as c:
        silinen = 0
        if truncate != "false" and ilce:
            silinen = c.execute(text("DELETE FROM uretim WHERE uretim_yili=:y AND ilce=:i"), {"y":final_yil,"i":ilce}).rowcount
        for s in range(0, len(rows), 500):
            c.execute(text("INSERT INTO uretim (uretim_yili,il,ilce,koy,urun,tarim_sekli,uretim_cesidi,ekili_alan) VALUES (:uretim_yili,:il,:ilce,:koy,:urun,:tarim_sekli,:uretim_cesidi,:ekili_alan)"), rows[s:s+500])
        sure = round(time.perf_counter()-t0,2)
        _log_import(c,file.filename,fh,final_yil,len(rows),silinen,sure)
    return {"ok":True,"ilce":ilce,"yil":final_yil,"eklenen":len(rows),"silinen":silinen,"sure_sn":sure}


# ═══════════════════════════════════════════════════════════
# HAYVANCIPLIK
# ═══════════════════════════════════════════════════════════
@app.get("/api/hayvancilik")
def list_hayvancilik(yil: int = Query(2025), ilce: Optional[str] = Query(None), koy: Optional[str] = Query(None), page: int = Query(1,ge=1), limit: int = Query(100,ge=1,le=50000)):
    clauses, params = ["uretim_yili=:yil"], {"yil":yil}
    if ilce: clauses.append("UPPER(ilce)=UPPER(:ilce)"); params["ilce"]=ilce
    if koy:  clauses.append("UPPER(koy) LIKE UPPER(:koy)"); params["koy"]=f"%{koy}%"
    where = "WHERE "+" AND ".join(clauses); offset=(page-1)*limit
    sql = f"SELECT ilce,koy,SUM(sigir)::int AS sigir,SUM(manda)::int AS manda,SUM(koyun)::int AS koyun,SUM(keci)::int AS keci,SUM(sigir_isletme)::int AS sigir_isletme,SUM(manda_isletme)::int AS manda_isletme,SUM(koyun_isletme)::int AS koyun_isletme,SUM(keci_isletme)::int AS keci_isletme,SUM(toplam_isletme)::int AS toplam_isletme FROM hayvancilik {where} GROUP BY ilce,koy ORDER BY ilce,koy LIMIT :limit OFFSET :offset"
    cnt = f"SELECT COUNT(*) FROM (SELECT koy FROM hayvancilik {where} GROUP BY ilce,koy) s"
    with engine.connect() as conn:
        rows  = conn.execute(text(sql),{**params,"limit":limit,"offset":offset}).mappings().all()
        total = int(conn.execute(text(cnt),params).scalar() or 0)
    return {"data":[dict(r) for r in rows],**_pagination(total,page,limit)}


@app.get("/api/hayvancilik/ozet")
def hayvancilik_ozet(yil: int = Query(2025), ilce: Optional[str] = Query(None)):
    where, params = build_where(yil, ilce, None, None, None, None)
    sql = f"SELECT SUM(sigir)::int AS sigir_toplam,SUM(manda)::int AS manda_toplam,SUM(koyun)::int AS koyun_toplam,SUM(keci)::int AS keci_toplam,SUM(sigir_isletme)::int AS sigir_isletme,SUM(manda_isletme)::int AS manda_isletme,SUM(koyun_isletme)::int AS koyun_isletme,SUM(keci_isletme)::int AS keci_isletme,SUM(toplam_isletme)::int AS toplam_isletme,COUNT(DISTINCT koy)::int AS koy_sayisi FROM hayvancilik {where}"
    with engine.connect() as conn:
        row = conn.execute(text(sql),params).mappings().fetchone()
    return dict(row) if row else {}


@app.delete("/api/hayvancilik/temizle")
def hayvancilik_temizle(yil: Optional[int] = Query(None), ilce: Optional[str] = Query(None)):
    with engine.begin() as conn:
        if yil and ilce: r = conn.execute(text("DELETE FROM hayvancilik WHERE uretim_yili=:y AND UPPER(ilce)=UPPER(:i)"),{"y":yil,"i":ilce})
        elif yil:        r = conn.execute(text("DELETE FROM hayvancilik WHERE uretim_yili=:y"),{"y":yil})
        else:            r = conn.execute(text("DELETE FROM hayvancilik"))
    return {"silinen":r.rowcount}


def parse_hayv_xls(content: bytes, ilce_adi: str, yil: int) -> list[dict]:
    cells = _read_xls_cells(content)
    if not cells: return []
    max_row = max(r for r,c in cells)
    header_row = col_koy = col_sigir = col_manda = col_koyun = col_keci = None
    for r in range(min(25, max_row)):
        strs = {c: str(v) for c,v in ((c2,cells.get((r,c2),'')) for c2 in range(25)) if isinstance(v,str)}
        if any('Köy' in v and 'Mahalle' in v for v in strs.values()) and any('Sığır' in v or 'Koyun' in v or 'Manda' in v or 'Keçi' in v for v in strs.values()):
            header_row = r
            for col,v in strs.items():
                if 'Köy' in v and 'Mahalle' in v: col_koy = col
                if 'Sığır' in v: col_sigir = col
                if 'Manda' in v: col_manda = col
                if 'Koyun' in v: col_koyun = col
                if 'Keçi'  in v: col_keci  = col
            break
    if header_row is None or col_koy is None: return []
    def num(r,col):
        if col is None: return 0
        try: return max(0, int(float(cells.get((r,col),0) or 0)))
        except: return 0
    koyler: dict = {}
    for r in range(header_row+1, max_row+1):
        koy = cells.get((r,col_koy),'')
        if not isinstance(koy,str) or not koy.strip(): continue
        koy = koy.strip().upper()
        si,ma,ko,ke = num(r,col_sigir),num(r,col_manda),num(r,col_koyun),num(r,col_keci)
        if koy not in koyler: koyler[koy] = dict(sigir=0,manda=0,koyun=0,keci=0,sigir_isletme=0,manda_isletme=0,koyun_isletme=0,keci_isletme=0,toplam_isletme=0)
        k = koyler[koy]
        k['sigir']+=si; k['manda']+=ma; k['koyun']+=ko; k['keci']+=ke
        if si>0: k['sigir_isletme']+=1
        if ma>0: k['manda_isletme']+=1
        if ko>0: k['koyun_isletme']+=1
        if ke>0: k['keci_isletme']+=1
        k['toplam_isletme']+=1
    return [{"uretim_yili":yil,"il":"BURDUR","ilce":ilce_adi.upper(),"koy":koy,**v} for koy,v in koyler.items()]


@app.post("/api/import/hayvancilik")
async def import_hayvancilik(file: UploadFile = File(...), yil: Optional[str] = Form(None), truncate: Optional[str] = Form("true")):
    _check_file(file, {"xls","xlsx","xlsm","ods"})
    if (file.filename or "").rsplit(".",1)[-1].lower() != "xls": raise HTTPException(422, ".xls formatı bekleniyor")
    fname = file.filename or ""
    ilce_adi = fname.split(".")[0].split("_")[0].strip().upper()
    final_yil = _yil_from(yil, fname)
    content = await file.read()
    fh = _sha256(content)
    with engine.connect() as _c:
        dup = _check_duplicate(_c, fh)
    if dup: raise HTTPException(409, f"Bu dosya daha önce yüklenmiş ({dup['dosya_adi']})")
    rows = parse_hayv_xls(content, ilce_adi, final_yil)
    if not rows: raise HTTPException(422, "Geçerli veri bulunamadı")
    t0 = time.perf_counter()
    with engine.begin() as conn:
        silinen = 0
        if truncate != "false":
            silinen = conn.execute(text("DELETE FROM hayvancilik WHERE uretim_yili=:y AND ilce=:i"),{"y":final_yil,"i":ilce_adi}).rowcount
        for s in range(0,len(rows),200):
            conn.execute(text("INSERT INTO hayvancilik (uretim_yili,il,ilce,koy,sigir,manda,koyun,keci,sigir_isletme,manda_isletme,koyun_isletme,keci_isletme,toplam_isletme) VALUES (:uretim_yili,:il,:ilce,:koy,:sigir,:manda,:koyun,:keci,:sigir_isletme,:manda_isletme,:koyun_isletme,:keci_isletme,:toplam_isletme)"),rows[s:s+200])
        sure = round(time.perf_counter()-t0,2)
        _log_import(conn,fname,fh,final_yil,len(rows),silinen,sure)
    return {"ok":True,"ilce":ilce_adi,"yil":final_yil,"koy_sayisi":len(rows),"silinen":silinen,"sure_sn":sure}


# ═══════════════════════════════════════════════════════════
# KOOPERATİF
# ═══════════════════════════════════════════════════════════
@app.get("/api/kooperatif")
def list_kooperatif(ilce: Optional[str]=Query(None), koop_turu: Optional[str]=Query(None), ara: Optional[str]=Query(None), page: int=Query(1,ge=1), limit: int=Query(200,ge=1,le=5000)):
    clauses, params = [], {}
    if ilce:      clauses.append("UPPER(ilce)=UPPER(:ilce)"); params["ilce"]=ilce
    if koop_turu: clauses.append("UPPER(koop_turu) LIKE UPPER(:kt)"); params["kt"]=f"%{koop_turu}%"
    if ara:       clauses.append("(UPPER(koy_belde) LIKE UPPER(:ara) OR UPPER(baskan) LIKE UPPER(:ara))"); params["ara"]=f"%{ara}%"
    where = ("WHERE "+" AND ".join(clauses)) if clauses else ""; offset=(page-1)*limit
    with engine.connect() as conn:
        rows  = conn.execute(text(f"SELECT * FROM kooperatif {where} ORDER BY ilce,koop_turu,koy_belde LIMIT :limit OFFSET :offset"),{**params,"limit":limit,"offset":offset}).mappings().all()
        total = int(conn.execute(text(f"SELECT COUNT(*) FROM kooperatif {where}"),params).scalar() or 0)
    return {"data":[dict(r) for r in rows],**_pagination(total,page,limit)}


@app.get("/api/kooperatif/ozet")
def kooperatif_ozet():
    with engine.connect() as conn:
        rows  = conn.execute(text("SELECT koop_turu,COUNT(*)::int AS sayi,COUNT(DISTINCT ilce)::int AS ilce_sayisi FROM kooperatif GROUP BY koop_turu ORDER BY sayi DESC")).mappings().all()
        total = int(conn.execute(text("SELECT COUNT(*)::int FROM kooperatif")).scalar() or 0)
    return {"data":[dict(r) for r in rows],"toplam":total}


@app.delete("/api/kooperatif/temizle")
def kooperatif_temizle():
    with engine.begin() as conn:
        r = conn.execute(text("DELETE FROM kooperatif"))
    return {"silinen":r.rowcount}


def parse_koop_xls(content: bytes) -> list[dict]:
    cells = _read_xls_cells(content)
    if not cells: return []
    max_row = max(r for r,c in cells)
    rows = []
    for r in range(2, max_row+1):
        ilce,koy,ktype,ortak,baskan,telefon = cells.get((r,1)),cells.get((r,2)),cells.get((r,3)),cells.get((r,4)),cells.get((r,5)),cells.get((r,6))
        if not isinstance(ilce,str) or not _clean(ilce): continue
        if not isinstance(koy,str)  or not _clean(koy):  continue
        if not isinstance(ktype,str): continue
        tel = ""
        if telefon:
            try: tel = str(int(float(telefon)))
            except: tel = _clean(telefon, 30)
        ortak_sayisi = None
        if ortak is not None:
            try: ortak_sayisi = int(float(ortak))
            except: pass
        rows.append({"ilce":_clean(ilce,60),"koy_belde":_clean(koy,120),"koop_turu":_clean(ktype,80),"ortak_sayisi":ortak_sayisi,"baskan":_clean(baskan,200),"telefon":tel})
    return rows


@app.post("/api/import/kooperatif")
async def import_kooperatif(file: UploadFile=File(...), truncate: Optional[str]=Form("true")):
    _check_file(file, {"xls","xlsx"})
    content = await file.read()
    fh = _sha256(content)
    with engine.connect() as _c:
        dup = _check_duplicate(_c, fh)
    if dup: raise HTTPException(409, f"Bu dosya daha önce yüklenmiş ({dup['dosya_adi']})")
    rows = parse_koop_xls(content)
    if not rows: raise HTTPException(422, "Geçerli veri bulunamadı")
    t0 = time.perf_counter()
    with engine.begin() as conn:
        silinen = 0
        if truncate != "false": silinen = conn.execute(text("DELETE FROM kooperatif")).rowcount
        for s in range(0,len(rows),200):
            conn.execute(text("INSERT INTO kooperatif (ilce,koy_belde,koop_turu,ortak_sayisi,baskan,telefon) VALUES (:ilce,:koy_belde,:koop_turu,:ortak_sayisi,:baskan,:telefon)"),rows[s:s+200])
        sure = round(time.perf_counter()-t0,2)
        _log_import(conn,file.filename,fh,None,len(rows),silinen,sure)
    return {"ok":True,"eklenen":len(rows),"silinen":silinen,"sure_sn":sure}


# ═══════════════════════════════════════════════════════════
# SÜT DESTEKLEME
# ═══════════════════════════════════════════════════════════
@app.get("/api/sut")
def list_sut(yil: Optional[int]=Query(None), donem: Optional[str]=Query(None), ilce: Optional[str]=Query(None), koy: Optional[str]=Query(None), sort_by: Optional[str]=Query("destek_tutari"), sort_dir: Optional[str]=Query("desc"), page: int=Query(1,ge=1), limit: int=Query(100,ge=1,le=50000)):
    clauses, params = [], {}
    if yil:   clauses.append("yil=:yil"); params["yil"]=yil
    if donem: clauses.append("donem=:donem"); params["donem"]=donem
    if ilce:  clauses.append("UPPER(ilce)=UPPER(:ilce)"); params["ilce"]=ilce
    if koy:   clauses.append("UPPER(koy) LIKE UPPER(:koy)"); params["koy"]=f"%{koy}%"
    where = ("WHERE "+" AND ".join(clauses)) if clauses else ""; offset=(page-1)*limit
    oc = sort_by if sort_by in {"yil","donem","il","ilce","koy","temel_sut_lt","destek_tutari"} else "destek_tutari"
    od = "DESC" if str(sort_dir or "").lower()=="desc" else "ASC"
    sql = f"SELECT il,ilce,koy,SUM(temel_sut_lt)::numeric AS temel_sut_lt,SUM(destek_tutari)::numeric AS destek_tutari,COUNT(*)::int AS uretici_sayisi FROM sut_destekleme {where} GROUP BY il,ilce,koy ORDER BY {oc} {od} LIMIT :limit OFFSET :offset"
    cnt = f"SELECT COUNT(*) FROM (SELECT koy FROM sut_destekleme {where} GROUP BY il,ilce,koy) s"
    with engine.connect() as conn:
        rows  = conn.execute(text(sql),{**params,"limit":limit,"offset":offset}).mappings().all()
        total = int(conn.execute(text(cnt),params).scalar() or 0)
    return {"data":[dict(r) for r in rows],**_pagination(total,page,limit)}


@app.get("/api/sut/ozet")
def sut_ozet(yil: Optional[int]=Query(None), ilce: Optional[str]=Query(None)):
    clauses, params = [], {}
    if yil:  clauses.append("yil=:yil"); params["yil"]=yil
    if ilce: clauses.append("UPPER(ilce)=UPPER(:ilce)"); params["ilce"]=ilce
    where = ("WHERE "+" AND ".join(clauses)) if clauses else ""
    sql = f"SELECT SUM(temel_sut_lt)::numeric AS toplam_sut_lt,SUM(destek_tutari)::numeric AS toplam_tutar,COUNT(*)::int AS uretici_sayisi,COUNT(DISTINCT koy)::int AS koy_sayisi,COUNT(DISTINCT ilce)::int AS ilce_sayisi FROM sut_destekleme {where}"
    with engine.connect() as conn:
        row = conn.execute(text(sql),params).mappings().fetchone()
    return dict(row) if row else {}


@app.get("/api/sut/donemler")
def sut_donemler():
    with engine.connect() as conn:
        rows = conn.execute(text("SELECT DISTINCT donem,yil FROM sut_destekleme ORDER BY yil DESC,donem DESC")).mappings().all()
    return {"data":[dict(r) for r in rows]}


@app.delete("/api/sut/temizle")
def sut_temizle(donem: Optional[str]=Query(None), yil: Optional[int]=Query(None)):
    with engine.begin() as conn:
        if donem:   r = conn.execute(text("DELETE FROM sut_destekleme WHERE donem=:d"),{"d":donem})
        elif yil:   r = conn.execute(text("DELETE FROM sut_destekleme WHERE yil=:y"),{"y":yil})
        else:       r = conn.execute(text("DELETE FROM sut_destekleme"))
    return {"silinen":r.rowcount}


def parse_sut_xlsx(content: bytes, donem: str, yil: int) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active; rows = []
    for row in ws.iter_rows(min_row=11, values_only=True):
        if not row[1] or not isinstance(row[1],(int,float)): continue
        koy = str(row[8] or '').strip().upper(); ilce = str(row[9] or '').strip().upper()
        if not koy or not ilce: continue
        rows.append({"donem":donem,"yil":yil,"il":str(row[10] or '').strip().upper(),"ilce":ilce,"koy":koy,"temel_sut_lt":round(float(row[12] or 0),2),"destek_tutari":round(float(row[25] or 0),2)})
    wb.close(); return rows


@app.post("/api/import/sut")
async def import_sut(donem: Optional[str]=Form(None), yil: Optional[str]=Form(None), truncate: Optional[str]=Form("false"), file: UploadFile=File(...)):
    _check_file(file, {"xlsx","xls","xlsm","ods"})
    donem_val = (donem or "").strip(); yil_val = (yil or "").strip()
    if not yil_val or not yil_val.isdigit():
        m = re.search(r"20[0-9]{2}", donem_val); yil_val = m.group(0) if m else ""
    if not donem_val: donem_val = yil_val or "Bilinmiyor"
    if not yil_val: raise HTTPException(422, "Yıl tespit edilemedi. Dönem alanına yıl ekleyin.")
    content = await file.read(); fh = _sha256(content)
    with engine.connect() as _c:
        dup = _check_duplicate(_c, fh)
    if dup: raise HTTPException(409, f"Bu dosya daha önce yüklenmiş ({dup['dosya_adi']})")
    rows = parse_sut_xlsx(content, donem_val, int(yil_val))
    if not rows: raise HTTPException(422, "Geçerli veri bulunamadı")
    t0 = time.perf_counter()
    with engine.begin() as conn:
        silinen = 0
        if truncate != "false": silinen = conn.execute(text("DELETE FROM sut_destekleme WHERE donem=:d"),{"d":donem_val}).rowcount
        for s in range(0,len(rows),500):
            conn.execute(text("INSERT INTO sut_destekleme (donem,yil,il,ilce,koy,temel_sut_lt,destek_tutari) VALUES (:donem,:yil,:il,:ilce,:koy,:temel_sut_lt,:destek_tutari)"),rows[s:s+500])
        sure = round(time.perf_counter()-t0,2)
        _log_import(conn,file.filename,fh,int(yil_val),len(rows),silinen,sure)
    return {"ok":True,"donem":donem_val,"yil":int(yil_val),"eklenen":len(rows),"silinen":silinen,"sure_sn":sure}


# ═══════════════════════════════════════════════════════════
# ÖZET DESTEK TABLOLARI
# ═══════════════════════════════════════════════════════════
def _parse_ozet_xls(content: bytes, key_field: str, sc: int = 1) -> list[dict]:
    cells = _read_xls_cells(content)
    if not cells: return []
    max_row = max(r for r,c in cells)
    yil_cols = _parse_yil_cols(cells, sc)
    if not yil_cols: return []
    rows = []
    for r in range(1, max_row+1):
        ad = str(cells.get((r,0),'') or '').strip()
        if not ad or any(t in ad.lower() for t in ('genel toplam','toplam')): continue
        ad = re.sub(r'^[0-9]+-\s*','',ad).strip()
        for col, yil in yil_cols.items():
            tutar = cells.get((r,col))
            try: tutar = round(float(tutar),2) if tutar is not None else 0.0
            except: tutar = 0.0
            rows.append({key_field: ad, "yil": yil, "tutar_tl": tutar})
    return rows

def _parse_fark_prim_xls(content: bytes) -> list[dict]:
    cells = _read_xls_cells(content)
    if not cells: return []
    max_row = max(r for r,c in cells)
    yil_cols = _parse_yil_cols(cells, 2, 10)
    if not yil_cols: return []
    ana: dict[str,int] = {}
    for r in range(1, max_row+1):
        a = str(cells.get((r,0),'') or '').strip()
        b = str(cells.get((r,1),'') or '').strip()
        if not a or 'genel toplam' in a.lower(): continue
        kat = re.sub(r'\s+Toplam.*$','', a.replace('\n',' '), flags=re.IGNORECASE).strip()
        if 'toplam' in a.lower() or not b or b == a.split('\n')[0].strip(): ana[kat] = r
    rows = []
    for kat, r in ana.items():
        for col, yil in yil_cols.items():
            tutar = cells.get((r,col))
            try: tutar = round(float(tutar),2) if tutar is not None else 0.0
            except: tutar = 0.0
            rows.append({"kategori":kat,"yil":yil,"tutar_tl":tutar})
    return rows

# GET / DELETE / POST endpoint factory
def _make_ozet_endpoints(route: str, tbl: str, kf: str, kl: str, parser):
    @app.get(f"/api/{route}")
    def _list(yil: Optional[int]=Query(None), limit: int=Query(9999,le=99999)):
        clauses, params = [], {}
        if yil: clauses.append("yil=:yil"); params["yil"]=yil
        where = ("WHERE "+" AND ".join(clauses)) if clauses else ""; params["limit"]=limit
        with engine.connect() as conn:
            rows = conn.execute(text(f"SELECT {kf},yil,tutar_tl FROM {tbl} {where} ORDER BY {kf} ASC,yil ASC LIMIT :limit"),params).mappings().all()
        return {"data":[dict(r) for r in rows],"total":len(rows)}

    @app.get(f"/api/{route}/ozet")
    def _ozet():
        with engine.connect() as conn:
            rows = conn.execute(text(f"SELECT yil,SUM(tutar_tl)::numeric AS toplam_tl,COUNT(DISTINCT {kf})::int AS {kl}_sayisi FROM {tbl} GROUP BY yil ORDER BY yil DESC")).mappings().all()
        return {"data":[dict(r) for r in rows]}

    @app.delete(f"/api/{route}/temizle")
    def _temizle(yil: Optional[int]=Query(None)):
        with engine.begin() as conn:
            r = conn.execute(text(f"DELETE FROM {tbl}"+(" WHERE yil=:y" if yil else ""), {"y":yil} if yil else {}))
        return {"silinen":r.rowcount}

    @app.post(f"/api/import/{route}")
    async def _import(truncate: Optional[str]=Form("false"), file: UploadFile=File(...)):
        _check_file(file, {"xls","xlsx","xlsm"})
        content = await file.read(); fh = _sha256(content)
        with engine.connect() as _c:
            dup = _check_duplicate(_c, fh)
        if dup: raise HTTPException(409, f"Bu dosya daha önce yüklenmiş ({dup['dosya_adi']})")
        rows = parser(content)
        if not rows: raise HTTPException(422, "Geçerli veri bulunamadı")
        t0 = time.perf_counter()
        with engine.begin() as conn:
            silinen = 0
            if truncate != "false": silinen = conn.execute(text(f"DELETE FROM {tbl}")).rowcount
            conn.execute(text(f"INSERT INTO {tbl} ({kf},yil,tutar_tl) VALUES (:{kf},:yil,:tutar_tl)"),rows)
            sure = round(time.perf_counter()-t0,2)
            _log_import(conn,file.filename,fh,None,len(rows),silinen,sure)
        return {"ok":True,"eklenen":len(rows),"silinen":silinen,"sure_sn":sure}

_make_ozet_endpoints("alan-bazli",        "alan_bazli_destek",  "destek_adi", "destek",   lambda c: _parse_ozet_xls(c,"destek_adi",1))
_make_ozet_endpoints("hayvancilik-destek","hayvancilik_destek", "destek_adi", "destek",   lambda c: _parse_ozet_xls(c,"destek_adi",1))
_make_ozet_endpoints("genel-destek",      "genel_destek",       "destek_adi", "destek",   lambda c: _parse_ozet_xls(c,"destek_adi",1))
_make_ozet_endpoints("fark-prim",         "fark_prim_destek",   "kategori",   "kategori", _parse_fark_prim_xls)


# ═══════════════════════════════════════════════════════════
# ÇKS SAYISI
# ═══════════════════════════════════════════════════════════
def _norm_koy(s: str) -> str:
    return re.sub(r'\s+(KÖYÜ?|MAHALLESİ?|BELDESİ?|MH\.?|KÖY\.?)$','',s.strip().upper()).strip()

def parse_cks_xlsx(content: bytes, yil: int) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True); ws = wb.active
    data_start = 1
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
        if row[0] and str(row[0]).strip().upper() in ('İLÇESİ','İLÇE ADI','İLÇE'):
            data_start = i+1; break
    rows = []
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        ilce=str(row[0] or '').strip(); koy=str(row[1] or '').strip(); sayi=row[2]
        if not ilce or not koy or sayi is None: continue
        try: sayi=int(float(sayi))
        except: continue
        rows.append({"yil":yil,"ilce":ilce.upper(),"koy":_norm_koy(koy),"sayi":sayi})
    wb.close(); return rows

@app.get("/api/cks-sayisi")
def list_cks(yil: Optional[int]=Query(None)):
    clauses, params = [], {}
    if yil: clauses.append("yil=:yil"); params["yil"]=yil
    where = ("WHERE "+" AND ".join(clauses)) if clauses else ""
    with engine.connect() as conn:
        rows = conn.execute(text(f"SELECT yil,ilce,koy,sayi FROM cks_sayisi {where} ORDER BY ilce,koy,yil"),params).mappings().all()
    return {"data":[dict(r) for r in rows],"total":len(rows)}

@app.get("/api/cks-sayisi/ozet")
def cks_ozet():
    with engine.connect() as conn:
        rows = conn.execute(text("SELECT yil,SUM(sayi)::int AS toplam_ciftci,COUNT(DISTINCT koy)::int AS koy_sayisi FROM cks_sayisi GROUP BY yil ORDER BY yil DESC")).mappings().all()
    return {"data":[dict(r) for r in rows]}

@app.delete("/api/cks-sayisi/temizle")
def cks_temizle(yil: Optional[int]=Query(None)):
    with engine.begin() as conn:
        r = conn.execute(text("DELETE FROM cks_sayisi"+(" WHERE yil=:y" if yil else ""),{"y":yil} if yil else {}))
    return {"silinen":r.rowcount}

@app.post("/api/import/cks-sayisi")
async def import_cks(yil: str=Form(...), truncate: Optional[str]=Form("false"), file: UploadFile=File(...)):
    if not yil.isdigit(): raise HTTPException(422,"Geçersiz yıl")
    _check_file(file, {"xlsx","xls","xlsm"})
    content = await file.read(); fh = _sha256(content)
    with engine.connect() as _c:
        dup = _check_duplicate(_c, fh)
    if dup: raise HTTPException(409, f"Bu dosya daha önce yüklenmiş ({dup['dosya_adi']})")
    rows = parse_cks_xlsx(content, int(yil))
    if not rows: raise HTTPException(422,"Geçerli veri bulunamadı")
    t0 = time.perf_counter()
    with engine.begin() as conn:
        silinen = 0
        if truncate != "false": silinen = conn.execute(text("DELETE FROM cks_sayisi WHERE yil=:y"),{"y":int(yil)}).rowcount
        conn.execute(text("INSERT INTO cks_sayisi (yil,ilce,koy,sayi) VALUES (:yil,:ilce,:koy,:sayi)"),rows)
        sure = round(time.perf_counter()-t0,2)
        _log_import(conn,file.filename,fh,int(yil),len(rows),silinen,sure)
    return {"ok":True,"yil":int(yil),"eklenen":len(rows),"silinen":silinen,"sure_sn":sure}


# ═══════════════════════════════════════════════════════════
# BİTKİSEL DESTEK
# ═══════════════════════════════════════════════════════════
def parse_bitkisel_destek_xls(content: bytes, yil: int) -> list[dict]:
    cells = _read_xls_cells(content)
    if not cells: return []
    max_row = max(r for r,c in cells); rows = []
    for r in range(4, max_row+1):
        ilce = _clean(cells.get((r,3)),60); koy = _clean(cells.get((r,4)),120); urun = _clean(cells.get((r,10)),120)
        if not ilce or not koy or not urun: continue
        rows.append({"yil":yil,"il":"BURDUR","ilce":ilce.upper(),"koy":koy.upper(),"urun":urun.upper(),
                     "feromon_adet":_num(cells.get((r,11))),"feromon_tuzak_adet":_num(cells.get((r,12))),
                     "faydali_bocek_adet":_num(cells.get((r,13))),"desteklenen_alan_da":_num(cells.get((r,14))),
                     "destek_tutari_tl":_num(cells.get((r,15))),"net_odeme_tl":_num(cells.get((r,17)))})
    return rows

@app.get("/api/bitkisel-destek")
def list_bitkisel(yil: int=Query(2025), ilce: Optional[str]=Query(None), koy: Optional[str]=Query(None), urun: Optional[str]=Query(None), sort_by: Optional[str]=Query("desteklenen_alan_da"), sort_dir: Optional[str]=Query("desc"), page: int=Query(1,ge=1), limit: int=Query(100,ge=1,le=50000)):
    clauses, params = ["yil=:yil"], {"yil":yil}
    if ilce: clauses.append("UPPER(ilce)=UPPER(:ilce)"); params["ilce"]=ilce
    if koy:  clauses.append("UPPER(koy) LIKE UPPER(:koy)"); params["koy"]=f"%{koy}%"
    if urun: clauses.append("UPPER(urun) LIKE UPPER(:urun)"); params["urun"]=f"%{urun}%"
    where = "WHERE "+" AND ".join(clauses); offset=(page-1)*limit
    _ALL = {"ilce","koy","urun","feromon_adet","feromon_tuzak_adet","faydali_bocek_adet","desteklenen_alan_da","destek_tutari_tl","net_odeme_tl"}
    oc = sort_by if sort_by in _ALL else "desteklenen_alan_da"
    od = "DESC" if str(sort_dir or "").lower()=="desc" else "ASC"
    sql = f"SELECT id,yil,il,ilce,koy,urun,feromon_adet,feromon_tuzak_adet,faydali_bocek_adet,desteklenen_alan_da,destek_tutari_tl,net_odeme_tl FROM bitkisel_destek {where} ORDER BY {oc} {od} LIMIT :limit OFFSET :offset"
    with engine.connect() as conn:
        rows  = conn.execute(text(sql),{**params,"limit":limit,"offset":offset}).mappings().all()
        total = int(conn.execute(text(f"SELECT COUNT(*) FROM bitkisel_destek {where}"),params).scalar() or 0)
    return {"data":[dict(r) for r in rows],**_pagination(total,page,limit)}

@app.get("/api/bitkisel-destek/ozet")
def bitkisel_ozet(yil: int=Query(2025), ilce: Optional[str]=Query(None), group_by: str=Query("ilce")):
    clauses, params = ["yil=:yil"], {"yil":yil}
    if ilce: clauses.append("UPPER(ilce)=UPPER(:ilce)"); params["ilce"]=ilce
    where = "WHERE "+" AND ".join(clauses)
    grp = group_by if group_by in {"ilce","koy","urun"} else "ilce"
    sql = f"SELECT {grp},COUNT(*)::int AS kayit_sayisi,ROUND(SUM(feromon_adet)::numeric,0) AS feromon_toplam,ROUND(SUM(feromon_tuzak_adet)::numeric,0) AS feromon_tuzak_toplam,ROUND(SUM(faydali_bocek_adet)::numeric,0) AS faydali_bocek_toplam,ROUND(SUM(desteklenen_alan_da)::numeric,2) AS alan_toplam,ROUND(SUM(destek_tutari_tl)::numeric,2) AS destek_toplam,ROUND(SUM(net_odeme_tl)::numeric,2) AS net_toplam FROM bitkisel_destek {where} GROUP BY {grp} ORDER BY destek_toplam DESC"
    with engine.connect() as conn:
        rows = conn.execute(text(sql),params).mappings().all()
    return {"group_by":grp,"data":[dict(r) for r in rows]}

@app.get("/api/bitkisel-destek/urunler")
def bitkisel_urunler(yil: int=Query(2025)):
    with engine.connect() as conn:
        rows = conn.execute(text("SELECT DISTINCT urun FROM bitkisel_destek WHERE yil=:yil ORDER BY urun"),{"yil":yil}).fetchall()
    return {"data":[r[0] for r in rows]}

@app.delete("/api/bitkisel-destek/temizle")
def bitkisel_temizle(yil: Optional[int]=Query(None)):
    with engine.begin() as conn:
        r = conn.execute(text("DELETE FROM bitkisel_destek"+(" WHERE yil=:y" if yil else ""),{"y":yil} if yil else {}))
    return {"silinen":r.rowcount}

@app.post("/api/import/bitkisel-destek")
async def import_bitkisel(file: UploadFile=File(...), yil: Optional[str]=Form(None), truncate: Optional[str]=Form("false")):
    _check_file(file, {"xls","xlsx","xlsm"})
    content = await file.read(); fh = _sha256(content)
    with engine.connect() as _c:
        dup = _check_duplicate(_c, fh)
    if dup: raise HTTPException(409, f"Bu dosya daha önce yüklenmiş ({dup['dosya_adi']})")
    final_yil = _yil_from(yil, file.filename or "")
    rows = parse_bitkisel_destek_xls(content, final_yil)
    if not rows: raise HTTPException(422,"Geçerli veri bulunamadı")
    t0 = time.perf_counter(); eklenen = guncellenen = 0
    with engine.begin() as conn:
        if truncate != "false": conn.execute(text("DELETE FROM bitkisel_destek WHERE yil=:y"),{"y":final_yil})
        for row in rows:
            res = conn.execute(text("""INSERT INTO bitkisel_destek (yil,il,ilce,koy,urun,feromon_adet,feromon_tuzak_adet,faydali_bocek_adet,desteklenen_alan_da,destek_tutari_tl,net_odeme_tl,updated_at) VALUES (:yil,:il,:ilce,:koy,:urun,:feromon_adet,:feromon_tuzak_adet,:faydali_bocek_adet,:desteklenen_alan_da,:destek_tutari_tl,:net_odeme_tl,NOW()) ON CONFLICT (yil,ilce,koy,urun) DO UPDATE SET feromon_adet=bitkisel_destek.feromon_adet+EXCLUDED.feromon_adet,feromon_tuzak_adet=bitkisel_destek.feromon_tuzak_adet+EXCLUDED.feromon_tuzak_adet,faydali_bocek_adet=bitkisel_destek.faydali_bocek_adet+EXCLUDED.faydali_bocek_adet,desteklenen_alan_da=bitkisel_destek.desteklenen_alan_da+EXCLUDED.desteklenen_alan_da,destek_tutari_tl=bitkisel_destek.destek_tutari_tl+EXCLUDED.destek_tutari_tl,net_odeme_tl=bitkisel_destek.net_odeme_tl+EXCLUDED.net_odeme_tl,updated_at=NOW()"""),row)
            if res.rowcount==1: eklenen+=1
            else: guncellenen+=1
        sure = round(time.perf_counter()-t0,2)
        _log_import(conn,file.filename,fh,final_yil,len(rows),0,sure)
    return {"ok":True,"yil":final_yil,"eklenen":eklenen,"guncellenen":guncellenen,"sure_sn":sure}


# ═══════════════════════════════════════════════════════════
# HEALTH
# ═══════════════════════════════════════════════════════════
@app.get("/health")
def health():
    with engine.connect() as c: c.execute(text("SELECT 1"))
    return {"status":"healthy"}
