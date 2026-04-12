#!/usr/bin/env python3
"""
ÇKS Excel Toplu Yükleme — Tek seferlik kullanım
================================================
pip install asyncpg openpyxl python-dotenv

Kullanım:
    # Tüm ilçe dosyalarını yükle
    python bulk_import.py --db "postgresql://user:pass@host/burdurdb" --dir ./excel

    # Tek dosya
    python bulk_import.py --db "..." --file AĞLASUN_CKS_....xlsx
"""
import argparse
import asyncio
import glob
import io
import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path

try:
    import asyncpg
except ImportError:
    sys.exit("❌  pip install asyncpg")
try:
    import openpyxl
except ImportError:
    sys.exit("❌  pip install openpyxl")

# Kolon indexleri (0-tabanlı)
COL_IL=3; COL_ILCE=4; COL_KOY=5; COL_URUN=11
COL_TARIM_SEKLI=12; COL_EKILI_ALAN=16; COL_URETIM_CESIDI=17
DATA_START=7


def parse_yil(ws) -> int:
    for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
        for cell in row:
            if cell and isinstance(cell, str) and "Üretim Yılı" in cell:
                m = re.search(r"(\d{4})", cell)
                if m: return int(m.group(1))
    return datetime.now().year


def read_excel(path: str) -> tuple[list[dict], int, str]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    yil  = parse_yil(ws)
    rows = []
    ilce = ""
    skipped = 0

    for row in ws.iter_rows(min_row=DATA_START, values_only=True):
        il_v=row[COL_IL]; ilce_v=row[COL_ILCE]; koy_v=row[COL_KOY]; urun_v=row[COL_URUN]
        if not il_v or not ilce_v or not koy_v or not urun_v:
            skipped += 1; continue
        try:    alan = float(row[COL_EKILI_ALAN] or 0)
        except: alan = 0.0
        if not ilce: ilce = str(ilce_v).strip().upper()
        rows.append({
            "yil": yil, "il": str(il_v).strip().upper(),
            "ilce": str(ilce_v).strip().upper(), "koy": str(koy_v).strip(),
            "urun": str(urun_v).strip(),
            "sekli":  str(row[COL_TARIM_SEKLI]   or "Kuru").strip(),
            "cesidi": str(row[COL_URETIM_CESIDI] or "1.Üretim").strip(),
            "alan": round(alan, 3),
        })

    wb.close()
    print(f"  📂  {Path(path).name}: {len(rows):,} satır okundu, {skipped} atlandı (yıl: {yil})")
    return rows, yil, ilce


async def import_file(pool: asyncpg.Pool, path: str, truncate=True):
    rows, yil, ilce = read_excel(path)
    if not rows:
        return 0, 0

    t0 = time.perf_counter()
    async with pool.acquire() as conn:
        async with conn.transaction():
            silinen = 0
            if truncate and ilce:
                result  = await conn.execute("DELETE FROM uretim WHERE uretim_yili=$1 AND ilce=$2", yil, ilce)
                try: silinen = int(result.split()[-1])
                except: pass
                if silinen:
                    print(f"  🗑️  {silinen:,} eski kayıt silindi ({ilce}/{yil})")

            BATCH = 500
            for s in range(0, len(rows), BATCH):
                chunk = rows[s:s+BATCH]
                await conn.executemany(
                    "INSERT INTO uretim (uretim_yili,il,ilce,koy,urun,tarim_sekli,uretim_cesidi,ekili_alan) VALUES($1,$2,$3,$4,$5,$6,$7,$8)",
                    [(r["yil"],r["il"],r["ilce"],r["koy"],r["urun"],r["sekli"],r["cesidi"],r["alan"]) for r in chunk],
                )
                print(f"  ↳  {min(s+BATCH,len(rows)):,}/{len(rows):,}", end="\r")
            print()

            sure = round(time.perf_counter()-t0, 2)
            await conn.execute(
                "INSERT INTO import_log(dosya_adi,ilce,uretim_yili,kayit_sayisi,silinen,sure_sn,durum) VALUES($1,$2,$3,$4,$5,$6,'basarili')",
                Path(path).name, ilce, yil, len(rows), silinen, sure,
            )

    print(f"  ✅  {len(rows):,} kayıt eklendi — {sure}s\n")
    return len(rows), silinen


async def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--db",          required=True)
    ap.add_argument("--dir",         default=None)
    ap.add_argument("--file",        default=None)
    ap.add_argument("--no-truncate", action="store_true")
    args = ap.parse_args()

    files = [args.file] if args.file else sorted(glob.glob(os.path.join(args.dir, "*CKS*.xlsx")))
    if not files:
        sys.exit("❌  Hiç dosya bulunamadı")

    print(f"\n🚀  {len(files)} dosya aktarılacak\n")
    pool = await asyncpg.create_pool(args.db, min_size=1, max_size=5)

    toplam = 0
    for i, f in enumerate(files, 1):
        print(f"[{i}/{len(files)}] {Path(f).name}")
        eklenen, _ = await import_file(pool, f, truncate=not args.no_truncate)
        toplam += eklenen

    await pool.close()
    print(f"═══════════════════\n✅  Toplam: {toplam:,} kayıt\n")

if __name__ == "__main__":
    asyncio.run(main())
