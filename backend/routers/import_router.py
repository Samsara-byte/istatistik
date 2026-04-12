"""
POST /api/import  —  ÇKS Excel dosyasını parse edip PostgreSQL'e yazar
"""
import io
import re
import time
from datetime import datetime
from typing import Optional
from fastapi import APIRouter, UploadFile, File, Form, HTTPException
import openpyxl
from database import get_pool

router = APIRouter(prefix="/api/import", tags=["import"])

# Excel kolon indexleri (0-tabanlı, iter_rows values_only=True)
COL_IL            = 3   # D — İl
COL_ILCE          = 4   # E — İlçe
COL_KOY           = 5   # F — Köy
COL_URUN          = 11  # L — Ürün
COL_TARIM_SEKLI   = 12  # M — Tarım Şekli
COL_EKILI_ALAN    = 16  # Q — Ekili Alan (da)
COL_URETIM_CESIDI = 17  # R — Üretim Çeşidi
DATA_START        = 7   # Veri satırı başlangıcı (Excel satır no, 1-tabanlı)


def _parse_yil(ws) -> int:
    """Başlık bölümünden 'Üretim Yılı: YYYY' yakala."""
    for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
        for cell in row:
            if cell and isinstance(cell, str) and "Üretim Yılı" in cell:
                m = re.search(r"(\d{4})", cell)
                if m:
                    return int(m.group(1))
    return datetime.now().year


def _parse_rows(ws, yil: int) -> tuple[list[dict], str, int]:
    """Worksheet satırlarını temizle, döndür."""
    rows: list[dict] = []
    skipped = 0
    ilce = ""

    for row in ws.iter_rows(min_row=DATA_START, values_only=True):
        il_val   = row[COL_IL]
        ilce_val = row[COL_ILCE]
        koy_val  = row[COL_KOY]
        urun_val = row[COL_URUN]

        if not il_val or not ilce_val or not koy_val or not urun_val:
            skipped += 1
            continue

        try:
            alan = float(row[COL_EKILI_ALAN] or 0)
        except (TypeError, ValueError):
            alan = 0.0

        sekli  = str(row[COL_TARIM_SEKLI]   or "Kuru").strip()
        cesidi = str(row[COL_URETIM_CESIDI] or "1.Üretim").strip()

        if not ilce:
            ilce = str(ilce_val).strip().upper()

        rows.append({
            "uretim_yili":   yil,
            "il":            str(il_val).strip().upper(),
            "ilce":          str(ilce_val).strip().upper(),
            "koy":           str(koy_val).strip(),
            "urun":          str(urun_val).strip(),
            "tarim_sekli":   sekli,
            "uretim_cesidi": cesidi,
            "ekili_alan":    round(alan, 3),
        })

    return rows, ilce, skipped


@router.post("")
async def import_excel(
    file:     UploadFile    = File(...),
    yil:      Optional[str] = Form(None),
    truncate: Optional[str] = Form("true"),
):
    # ── Uzantı kontrolü ────────────────────────────────────────
    ext = (file.filename or "").rsplit(".", 1)[-1].lower()
    if ext not in {"xlsx", "xls", "xlsm", "ods"}:
        raise HTTPException(400, "Sadece .xlsx / .xls / .ods dosyaları kabul edilir")

    content = await file.read()
    if len(content) > 60 * 1024 * 1024:
        raise HTTPException(413, "Dosya çok büyük (maks 60 MB)")

    # ── Excel parse ────────────────────────────────────────────
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(422, f"Excel okunamadı: {e}")

    file_yil  = _parse_yil(ws)
    final_yil = int(yil) if (yil and yil.isdigit()) else file_yil
    rows, ilce, skipped = _parse_rows(ws, final_yil)
    wb.close()

    if not rows:
        raise HTTPException(422, f"Geçerli veri bulunamadı (atlanan: {skipped} satır)")

    do_truncate = (truncate != "false")
    t0 = time.perf_counter()

    # ── DB insert ──────────────────────────────────────────────
    pool = get_pool()
    async with pool.acquire() as conn:
        async with conn.transaction():

            # Eski kayıtları sil
            silinen = 0
            if do_truncate and ilce:
                result  = await conn.execute(
                    "DELETE FROM uretim WHERE uretim_yili=$1 AND ilce=$2",
                    final_yil, ilce,
                )
                try:
                    silinen = int(result.split()[-1])
                except Exception:
                    silinen = 0

            # Toplu insert (500'lük batch)
            BATCH = 500
            for start in range(0, len(rows), BATCH):
                chunk = rows[start:start + BATCH]
                await conn.executemany(
                    """
                    INSERT INTO uretim
                        (uretim_yili, il, ilce, koy, urun,
                         tarim_sekli, uretim_cesidi, ekili_alan)
                    VALUES ($1,$2,$3,$4,$5,$6,$7,$8)
                    """,
                    [(r["uretim_yili"], r["il"], r["ilce"], r["koy"], r["urun"],
                      r["tarim_sekli"], r["uretim_cesidi"], r["ekili_alan"])
                     for r in chunk],
                )

            # Import log
            sure = round(time.perf_counter() - t0, 2)
            await conn.execute(
                """
                INSERT INTO import_log
                    (dosya_adi, ilce, uretim_yili, kayit_sayisi,
                     silinen, sure_sn, durum)
                VALUES ($1,$2,$3,$4,$5,$6,'basarili')
                """,
                file.filename, ilce, final_yil,
                len(rows), silinen, sure,
            )

    return {
        "ok":      True,
        "ilce":    ilce,
        "yil":     final_yil,
        "eklenen": len(rows),
        "silinen": silinen,
        "atlandi": skipped,
        "sure_sn": sure,
    }
